"""Read-only redacted summary of a pilot output directory.

Designed for use during a real-file pilot, where the operator must
NEVER paste raw business content into a chat, ticket, or commit. This
module walks the standard learn → confirm-mapping → run-preview →
render-docx artifacts under ``--out`` and reports a per-stage
**aggregate** picture: file presence and size, integer counts already
present in each YAML/XLSX ``summary`` block or ``Status`` column, and
a single next-action hint per stage.

The redaction contract is about what reaches the **printed output**,
not about what touches Python memory while parsing. Specifically:

  * **Load surface.** Two artifacts are parsed by their standard
    libraries: ``auto_mapping.yml`` / ``confirmed_mapping.yml`` /
    ``render_log.yml`` via PyYAML's ``safe_load`` (which materialises
    the full document as a Python dict), and ``run_validation.xlsx``
    via ``openpyxl.load_workbook(read_only=True)`` plus
    ``iter_rows(values_only=True)`` (which yields a tuple of EVERY
    cell value per row — Word Raw Token, Source Sheet, Source Cell,
    Generated Value, Detail, the lot — and we extract only the
    ``Status`` cell into the ``counts`` dict). Sensitive non-summary
    YAML fields and non-Status XLSX cell values therefore transit
    Python memory during parsing. We do not attempt to characterise
    when those PyYAML / openpyxl internal objects (the parsed YAML
    document, openpyxl's shared-string and sheet caches, etc.)
    become eligible for garbage collection — that is library
    implementation detail and outside our control. ``wb.close()``
    runs in a ``finally`` block to release the underlying file handle
    promptly, but releasing Python objects is the GC's job, not
    ``close()``'s. The redaction contract is enforced at the next
    two surfaces, not at this one. The remaining artifacts —
    ``mapping_review.xlsx``, ``confidence_report.md``,
    ``converted_template.docx``, and any rendered ``.docx`` — are
    detected by ``Path.exists`` / ``Path.stat`` / ``Path.iterdir`` /
    suffix check only; their contents are never opened, parsed, or
    held in memory by this module.
  * **Extract surface.** Only aggregate fields and policy enum
    strings reach :class:`StageStatus`: integer counts from each
    artifact's ``summary`` block, distinct ``Status`` values from
    ``run_validation.xlsx``, the boolean flags ``summary.complete`` /
    ``summary.allow_incomplete``, file sizes, and presence markers.
    Nothing else is stored on the report object.
  * **Print surface.** :func:`format_summary` reads only fields on
    :class:`StageStatus` plus :attr:`PilotSummary.out_basename`. By
    construction it cannot print raw Word tokens, generated values,
    raw Excel values, source sheet/cell addresses, Word location
    strings, Word snippet/context, reviewer notes, ``Detail``
    messages, ``source_artifacts`` / ``inputs`` file paths,
    ``out_docx`` paths, or any individual ``word_id``. The test
    suite pins this by planting unique sentinel strings in every
    content field of every artifact and asserting that none of them
    surface in the formatter output.

Filenames are reported as basenames; the ``--out`` directory is
reduced to its basename so a copy-pasted summary does not leak the
operator's home-directory layout.

Failure surface (the CLI translates these to exit codes):

  * The pre-condition ``--out`` exists is enforced by the CLI (exit 2).
  * If ``--out`` exists but the minimum required artifact
    (``auto_mapping.yml``) is absent, ``summarize_pilot`` records a
    fatal error and the CLI exits 11. There is nothing to summarize.

Read-only by design: never re-runs the pipeline, never mutates an
artifact, never calls out to an LLM, GUI, network, or Microsoft Word
automation. Matching, confirmation, rendering, and validation logic
are untouched.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import yaml


# Standard learn-mode artifacts the pipeline writes alongside one
# another under ``--out``. Order matters only for stable presentation —
# the validator does not depend on it.
_LEARN_ARTIFACTS: Tuple[str, ...] = (
    "auto_mapping.yml",
    "mapping_review.xlsx",
    "confidence_report.md",
    "converted_template.docx",
)
# The minimum required artifact for a summary to be meaningful. If the
# operator has not run ``learn`` yet there is no audit trail to inspect.
_REQUIRED_ARTIFACT = "auto_mapping.yml"

_CONFIRMED_FILE = "confirmed_mapping.yml"
# ``run-preview`` writes its artifact under a subdirectory of ``--out``
# in the canonical pilot layout (``$PILOT/output/run_preview/``); we
# also accept it at the top level for operators who pointed
# ``run-preview --out`` directly at ``$PILOT/output``.
_RUN_VALIDATION_BASENAME = "run_validation.xlsx"
_RUN_VALIDATION_SUBDIR = "run_preview"

_RENDER_LOG = "render_log.yml"
# Anything named other than the template counts as a candidate rendered
# docx — we only count, never name, so an operator-chosen filename
# never lands in the summary text.
_TEMPLATE_DOCX = "converted_template.docx"


# ---------------------------------------------------------------------------
# Public dataclasses
# ---------------------------------------------------------------------------

@dataclass
class StageStatus:
    """One pipeline stage's worth of redacted observations.

    ``details`` is an ordered list of ``(label, value)`` pairs the
    formatter renders verbatim. Keeping it as opaque strings makes the
    redaction contract simple: producers only insert counts and policy
    enum values, so the formatter cannot accidentally print a raw
    metric simply by changing template syntax.
    """
    name: str
    present: bool = False
    details: List[Tuple[str, str]] = field(default_factory=list)
    issues: List[str] = field(default_factory=list)
    next_action: Optional[str] = None


@dataclass
class PilotSummary:
    """The four-stage view a reviewer sees on the CLI."""
    out_basename: str
    learn: StageStatus = field(default_factory=lambda: StageStatus(name="learn"))
    confirm: StageStatus = field(
        default_factory=lambda: StageStatus(name="confirm-mapping")
    )
    run_preview: StageStatus = field(
        default_factory=lambda: StageStatus(name="run-preview")
    )
    render: StageStatus = field(
        default_factory=lambda: StageStatus(name="render-docx")
    )
    fatal_errors: List[str] = field(default_factory=list)

    @property
    def stages(self) -> List[StageStatus]:
        return [self.learn, self.confirm, self.run_preview, self.render]

    @property
    def ok(self) -> bool:
        return not self.fatal_errors


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def summarize_pilot(out_dir: Path) -> PilotSummary:
    """Inspect ``out_dir`` and build a redacted :class:`PilotSummary`.

    Pre-condition: ``out_dir`` exists (the CLI guards this for a
    friendlier error message). The function tolerates a partial pilot:
    only the learn-mode artifacts are required, and downstream stages
    are reported as ``[not yet run]`` rather than failing.
    """
    summary = PilotSummary(out_basename=_redact_dir_name(out_dir))

    auto_path = out_dir / _REQUIRED_ARTIFACT
    if not auto_path.exists():
        # Fail-clear: there is no audit trail to summarize. We
        # intentionally name the required artifact (a public filename)
        # rather than echoing back the full ``--out`` path, which could
        # contain a sensitive project codename.
        summary.fatal_errors.append(
            f"required artifact {_REQUIRED_ARTIFACT!s} not present in "
            f"{summary.out_basename!r} — run `learn` first."
        )
        return summary

    _summarize_learn(summary.learn, out_dir)
    _summarize_confirm(summary.confirm, out_dir)
    _summarize_run_preview(summary.run_preview, out_dir)
    _summarize_render(summary.render, out_dir)
    return summary


def format_summary(summary: PilotSummary) -> str:
    """Render a :class:`PilotSummary` as plain text safe to share.

    The formatter never reads beyond the ``StageStatus`` fields, so the
    redaction guarantee is enforced upstream (in :func:`summarize_pilot`).
    Output uses ASCII separators only — no emojis, no shell colour
    codes — so a pasted summary renders identically in a terminal, a
    chat client, or a ticket comment.
    """
    title = f"pilot-summary for: {summary.out_basename}"
    lines = [title, "=" * len(title), ""]

    if summary.fatal_errors:
        lines.append("FATAL:")
        for err in summary.fatal_errors:
            lines.append(f"  - {err}")
        lines.append("")
        lines.append(
            "No further stages inspected. Run `learn` against your pilot "
            "inputs first (see docs/real_file_pilot.md §2)."
        )
        return "\n".join(lines)

    for stage in summary.stages:
        marker = "[present]" if stage.present else "[not yet run]"
        lines.append(f"[{stage.name}] {marker}")
        for label, value in stage.details:
            lines.append(f"  {label:32s} : {value}")
        for issue in stage.issues:
            lines.append(f"  ! {issue}")
        if stage.next_action:
            lines.append(f"  -> next: {stage.next_action}")
        lines.append("")

    # Each stage prints its own next-action; a separate bottom-line
    # recommendation would either duplicate that or, worse, disagree
    # (e.g. claim "run-preview" when confirm-mapping is incomplete).
    return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------------
# Per-stage summaries — each writes ONLY counts / policy enums into the
# StageStatus. None reads raw Word/Excel content.
# ---------------------------------------------------------------------------

def _summarize_learn(stage: StageStatus, out_dir: Path) -> None:
    stage.present = True  # caller verified the required artifact exists

    for name in _LEARN_ARTIFACTS:
        p = out_dir / name
        if p.exists():
            stage.details.append((name, _safe_size(p)))
        else:
            stage.details.append((name, "MISSING"))
            stage.issues.append(
                f"{name} expected after `learn`, but not present"
            )

    auto_doc = _safe_load_yaml(out_dir / _REQUIRED_ARTIFACT, stage)
    if not isinstance(auto_doc, dict):
        # _safe_load_yaml already recorded the issue.
        return
    yaml_summary = auto_doc.get("summary") if isinstance(auto_doc.get("summary"), dict) else {}
    by_conf = yaml_summary.get("by_confidence") if isinstance(yaml_summary.get("by_confidence"), dict) else {}

    total = _safe_int(yaml_summary.get("total"))
    placeholders_applied = _safe_int(yaml_summary.get("placeholders_applied"))
    high = _safe_int(by_conf.get("HIGH"))
    medium = _safe_int(by_conf.get("MEDIUM"))
    low = _safe_int(by_conf.get("LOW"))
    unresolved = _safe_int(by_conf.get("UNRESOLVED"))
    excluded = _safe_int(by_conf.get("EXCLUDED"))

    stage.details.append(("total Word numbers", str(total)))
    stage.details.append(("  HIGH", str(high)))
    stage.details.append(("  MEDIUM", str(medium)))
    stage.details.append(("  LOW", str(low)))
    stage.details.append(("  UNRESOLVED", str(unresolved)))
    stage.details.append(("  EXCLUDED", str(excluded)))
    stage.details.append(("placeholders applied", str(placeholders_applied)))

    strict_blockers = low + unresolved
    if strict_blockers > 0:
        stage.next_action = (
            f"`learn --strict` would fail: {strict_blockers} eligible row(s) "
            "are LOW or UNRESOLVED. Review mapping_review.xlsx and "
            "confidence_report.md, then either fix the matcher gap or mark "
            "each row Reviewer Decision=reject."
        )
    else:
        stage.next_action = (
            "review mapping_review.xlsx (Reviewer Decision column) and run "
            "`confirm-mapping`."
        )


def _summarize_confirm(stage: StageStatus, out_dir: Path) -> None:
    path = out_dir / _CONFIRMED_FILE
    if not path.exists():
        stage.details.append((_CONFIRMED_FILE, "MISSING"))
        stage.next_action = (
            "fill in Reviewer Decision in mapping_review.xlsx and run "
            "`confirm-mapping`."
        )
        return
    stage.present = True
    stage.details.append((_CONFIRMED_FILE, _safe_size(path)))

    doc = _safe_load_yaml(path, stage)
    if not isinstance(doc, dict):
        return
    s = doc.get("summary") if isinstance(doc.get("summary"), dict) else {}

    complete = s.get("complete")
    allow_incomplete = s.get("allow_incomplete")
    confirmed_count = _safe_int(s.get("confirmed"))
    review_required_count = _safe_int(s.get("review_required"))
    audit_only_count = _safe_int(s.get("audit_only_excluded"))
    total = _safe_int(s.get("total_word_numbers"))

    stage.details.append(("summary.complete", _safe_bool_str(complete)))
    stage.details.append(("summary.allow_incomplete", _safe_bool_str(allow_incomplete)))
    stage.details.append(("confirmed", str(confirmed_count)))
    stage.details.append(("review_required", str(review_required_count)))
    stage.details.append(("audit_only_excluded", str(audit_only_count)))
    stage.details.append(("total Word numbers (recorded)", str(total)))

    if review_required_count > 0:
        stage.next_action = (
            f"`confirm-mapping` is incomplete: {review_required_count} row(s) "
            "still need reviewer action. Fix Reviewer Decision rows in "
            "mapping_review.xlsx and re-run."
        )
    elif allow_incomplete is True:
        # An exploratory --allow-incomplete run is never safe for a real
        # pilot; surface that loudly even when review_required is 0,
        # because some operator might have skipped a row by hand.
        stage.next_action = (
            "summary.allow_incomplete=true — this file was written with the "
            "exploratory escape hatch and must NOT be used for a real-file "
            "pilot. Re-run `confirm-mapping` without --allow-incomplete."
        )
    elif complete is True and confirmed_count > 0:
        stage.next_action = "run `run-preview` against the new-period workbook."
    else:
        # Defensive: complete is not True but review_required is 0 — the
        # YAML is in an unexpected shape. Surface so the operator
        # investigates rather than trusting the file silently.
        stage.next_action = (
            "confirmed_mapping.yml summary is incomplete-but-empty; inspect "
            "the file before continuing."
        )


def _summarize_run_preview(stage: StageStatus, out_dir: Path) -> None:
    path = _locate_run_validation(out_dir)
    if path is None:
        stage.details.append((_RUN_VALIDATION_BASENAME, "MISSING"))
        stage.next_action = (
            "run `run-preview` against the new-period workbook to produce "
            f"{_RUN_VALIDATION_BASENAME}."
        )
        return
    stage.present = True
    # Show the relative location so an operator who used a non-canonical
    # ``--out`` can tell whether the file is at the expected subdirectory.
    try:
        rel = path.relative_to(out_dir)
    except ValueError:
        rel = Path(path.name)
    stage.details.append((str(rel).replace("\\", "/"), _safe_size(path)))

    try:
        status_counts, total_rows = _count_status_column(path)
    except Exception as exc:  # broad on purpose: any read error gets surfaced
        stage.issues.append(
            f"{_RUN_VALIDATION_BASENAME}: cannot read ({exc.__class__.__name__})"
        )
        return
    stage.details.append(("total rows", str(total_rows)))
    ok_count = status_counts.get("ok", 0)
    stage.details.append(("  ok", str(ok_count)))
    # Only emit per-status counts for non-ok values so a clean run shows
    # one line and a broken run lists exactly what failed (status enum
    # strings only — never values, never cell addresses).
    non_ok = sorted(k for k in status_counts if k != "ok")
    for k in non_ok:
        stage.details.append((f"  {k}", str(status_counts[k])))

    failed = total_rows - ok_count
    if failed > 0:
        stage.next_action = (
            f"{failed} run-preview row(s) failed. Inspect run_validation.xlsx "
            "and re-run `run-preview` after fixing the new-period workbook "
            "or re-running `learn` + `confirm-mapping` on a fresh pair."
        )
    elif total_rows == 0:
        stage.next_action = (
            "run_validation.xlsx has no data rows. Re-run `run-preview` "
            "after the confirmed mapping has at least one confirmed entry."
        )
    else:
        stage.next_action = "run `render-docx` to produce the new Word report."


def _summarize_render(stage: StageStatus, out_dir: Path) -> None:
    log_path = out_dir / _RENDER_LOG
    rendered_docx_count = _count_rendered_docx(out_dir)

    if not log_path.exists() and rendered_docx_count == 0:
        stage.details.append((_RENDER_LOG, "MISSING"))
        stage.details.append(("rendered docx files", "0"))
        stage.next_action = (
            "run `render-docx` to produce the new Word report and its "
            f"{_RENDER_LOG}."
        )
        return

    stage.present = True
    if log_path.exists():
        stage.details.append((_RENDER_LOG, _safe_size(log_path)))
    else:
        stage.details.append((_RENDER_LOG, "MISSING"))
        stage.issues.append(
            f"{_RENDER_LOG} not present even though a rendered .docx is "
            "in the output directory; render audit is incomplete."
        )
    stage.details.append(("rendered docx files", str(rendered_docx_count)))

    doc = _safe_load_yaml(log_path, stage) if log_path.exists() else None
    if isinstance(doc, dict):
        s = doc.get("summary") if isinstance(doc.get("summary"), dict) else {}
        total_rows = _safe_int(s.get("total_rows"))
        ok = _safe_int(s.get("ok"))
        failed = _safe_int(s.get("failed"))
        total_replacements = _safe_int(s.get("total_replacements"))
        distinct = _safe_int(s.get("distinct_placeholder_word_ids"))
        stage.details.append(("log rows", str(total_rows)))
        stage.details.append(("  ok", str(ok)))
        stage.details.append(("  failed", str(failed)))
        stage.details.append(("total placeholder replacements", str(total_replacements)))
        stage.details.append(("distinct placeholder word_ids", str(distinct)))
        if failed > 0:
            stage.next_action = (
                f"{failed} render-docx row(s) failed. Re-run `render-docx` "
                "after fixing the upstream artifact (run_validation.xlsx or "
                "converted_template.docx)."
            )
        elif total_rows == 0:
            stage.next_action = (
                "render_log.yml has no replacement rows. Re-run the pipeline "
                "with a non-empty confirmed mapping."
            )
        else:
            stage.next_action = (
                "run `validate-render` to cross-check the rendered docx, "
                "render_log.yml, and run_validation.xlsx tell the same story."
            )
    else:
        stage.next_action = (
            "render_log.yml could not be read; re-run `render-docx` and then "
            "`validate-render`."
        )


# ---------------------------------------------------------------------------
# Locators & safe readers
# ---------------------------------------------------------------------------

def _locate_run_validation(out_dir: Path) -> Optional[Path]:
    """Find ``run_validation.xlsx`` under ``out_dir``.

    Canonical layout: ``out_dir/run_preview/run_validation.xlsx``. Some
    operators run ``run-preview --out`` pointing at the same directory
    as the rest of the pilot output; fall back to ``out_dir/run_validation.xlsx``
    in that case so a non-canonical layout still summarizes.
    """
    canonical = out_dir / _RUN_VALIDATION_SUBDIR / _RUN_VALIDATION_BASENAME
    if canonical.exists():
        return canonical
    flat = out_dir / _RUN_VALIDATION_BASENAME
    if flat.exists():
        return flat
    return None


def _count_status_column(path: Path) -> Tuple[Dict[str, int], int]:
    """Open ``path`` and count distinct ``Status`` column values.

    Load surface (called out so the module docstring's claim stays
    honest): ``iter_rows(values_only=True)`` yields a tuple per row
    that contains EVERY cell value — Word Raw Token, Source Sheet,
    Source Cell, Generated Value, Detail, the lot. We extract only
    ``raw[status_col]`` into ``counts``; the other cell values are
    read by openpyxl and end up in Python objects this function
    never inspects. ``wb.close()`` in the ``finally`` block releases
    the underlying file handle; when the matching Python objects
    (openpyxl's internal caches, individual cell values, the row
    tuples themselves) become eligible for garbage collection is a
    PyYAML/openpyxl/CPython concern we deliberately do not
    characterise here. None of the non-Status content reaches
    ``counts`` or, therefore, :class:`StageStatus` or the formatter;
    the redaction contract is about what the formatter prints, not
    about library memory lifetimes.
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return {}, 0
        try:
            status_col = header.index("Status")
        except ValueError:
            raise ValueError(
                f"{path.name} missing required Status column"
            )
        counts: Dict[str, int] = {}
        total = 0
        for raw in rows_iter:
            if raw is None:
                continue
            if all(cell is None or cell == "" for cell in raw):
                continue
            if status_col >= len(raw):
                continue
            v = raw[status_col]
            label = ("" if v is None else str(v)).strip() or "(empty)"
            counts[label] = counts.get(label, 0) + 1
            total += 1
        return counts, total
    finally:
        wb.close()


def _count_rendered_docx(out_dir: Path) -> int:
    """Count ``.docx`` files in ``out_dir`` other than the template.

    Filenames are never emitted — the operator already named them and
    listing them in a shared summary would leak chosen project/period
    labels. We only need the count to signal that the renderer ran.
    """
    if not out_dir.is_dir():
        return 0
    count = 0
    for p in out_dir.iterdir():
        if not p.is_file():
            continue
        if p.suffix.lower() != ".docx":
            continue
        if p.name == _TEMPLATE_DOCX:
            continue
        count += 1
    return count


def _safe_load_yaml(path: Path, stage: StageStatus):
    try:
        return yaml.safe_load(path.read_text(encoding="utf-8"))
    except Exception as exc:
        stage.issues.append(
            f"{path.name}: cannot parse ({exc.__class__.__name__})"
        )
        return None


def _safe_size(path: Path) -> str:
    try:
        return f"{path.stat().st_size:,} bytes"
    except OSError:
        return "present (size unreadable)"


def _safe_int(v) -> int:
    """Coerce a YAML/XLSX summary count to int without surfacing the source.

    Returns 0 on any unexpected shape; a producer who writes a string
    like ``"34"`` gets a meaningful number, while malformed values
    don't crash the summary.
    """
    if isinstance(v, bool):
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0
        try:
            return int(s)
        except ValueError:
            try:
                return int(float(s))
            except ValueError:
                return 0
    return 0


def _safe_bool_str(v) -> str:
    """Render a YAML bool as the literal text the policy contract uses.

    Uses ``true``/``false`` to match what ``confirmed_mapping.yml``
    serializes, so a reviewer reading the summary can match the words
    directly. Non-bool values are reported as ``(missing)`` rather than
    coerced so a malformed file is visible, not silently normalized.
    """
    if v is True:
        return "true"
    if v is False:
        return "false"
    return "(missing)"


def _redact_dir_name(out_dir: Path) -> str:
    """Reduce a path to its basename for safe sharing.

    The full ``--out`` path may contain the operator's home directory,
    a project codename, or a customer name. We keep just the leaf so a
    pasted summary has enough context to identify the run without
    leaking the parent tree.
    """
    name = out_dir.name
    if name:
        return name
    # ``Path("/")`` has empty ``.name``; fall back to a fixed label
    # rather than re-emitting the absolute path.
    return "(root)"
