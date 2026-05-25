"""Persist matcher output as human-reviewable artifacts.

Two outputs:
  * ``mapping_review.xlsx`` — one row per Word number with the top candidate
    and enough metadata for a reviewer to either confirm the mapping or
    correct it. This file is the substrate for the future human-confirmed
    mapping workflow.
  * ``confidence_report.md`` — narrative summary highlighting unresolved and
    low-confidence numbers, so the *first* thing a reviewer sees is what
    needs attention, not what already worked.

The workbook ships a second ``Summary`` sheet at tab position 0 so an
operator-pilot reviewer sees aggregate counts and next-action guidance
before they wade into the per-row details. ``mapping_review`` is kept as
the *active* sheet so downstream tools that read ``wb.active``
(``confirm-mapping``, ``validate-artifacts``, ``pilot-summary``)
continue to see the detail rows unchanged.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .template_builder import derive_review_status
from .validator import CoverageSummary
from .value_matcher import WordMatch

# The first four columns are the trust-slice keys: a stable Word ID (the
# join key with auto_mapping.yml) plus the three columns that record what
# the template builder did with the token. Keeping them up front means a
# reviewer skimming the XLSX sees the audit story before the candidate
# details.
#
# The final four columns are the reviewer-facing decision columns. They
# are written blank on every learn run and only `confirm-mapping` reads
# them back — never the matcher. Blank must never be treated as
# "confirmed"; that is enforced in src.mapping_confirmer.
REVIEW_HEADERS = [
    "Word ID",
    "Word Location",
    "Word Snippet",
    "Word Raw Token",
    "Word Value",
    "Word Unit",
    "Confidence",
    "Review Status",
    "Placeholder Status",
    "Placeholder",
    "Top Excel Sheet",
    "Top Excel Cell",
    "Top Excel Value",
    "Top Row Context",
    "Top Column Context",
    "Interpretation",
    "Value Score",
    "Context Score",
    "Overlap Tokens",
    "# Alt Candidates",
    "Note",
    "Reviewer Decision",
    "Reviewer Notes",
    "Confirmed Sheet",
    "Confirmed Cell",
]

REVIEWER_COLUMNS = ("Reviewer Decision", "Reviewer Notes", "Confirmed Sheet", "Confirmed Cell")

SUMMARY_SHEET_TITLE = "Summary"

# Confidence buckets shown on the Summary sheet, in the order a reviewer
# should walk them: HIGH first (quick confirms), then MEDIUM (needs eyes
# but recoverable), then LOW/UNRESOLVED (the audit risks), finally
# EXCLUDED (policy-only, never confirmed).
CONFIDENCE_ORDER = ("HIGH", "MEDIUM", "LOW", "UNRESOLVED", "EXCLUDED")

# Next-action guidance per confidence label. Mirrors the README's
# "Confidence statuses" table — kept here so the workbook explains itself
# without forcing a reviewer to alt-tab to the docs.
CONFIDENCE_GUIDANCE = {
    "HIGH": (
        "Safe to confirm — value+context agree. "
        "Tick `confirm` in Reviewer Decision (leave override columns blank) "
        "to accept the recommended source."
    ),
    "MEDIUM": (
        "Needs review — value+context agree but loosely or with a tied "
        "runner-up. Eyeball the Top Excel cell vs. the Word snippet before "
        "ticking `confirm`."
    ),
    "LOW": (
        "NOT safe to confirm — context overlap is weak. Investigate "
        "manually; leave Reviewer Decision blank to route the row into "
        "review_required."
    ),
    "UNRESOLVED": (
        "No Excel source — leave Reviewer Decision blank. Do not invent a "
        "source in the override columns; the row will land in "
        "review_required for follow-up."
    ),
    "EXCLUDED": (
        "Audit-only — date/period marker the matcher deliberately skipped. "
        "Even `confirm` keeps the row in audit_only_excluded; the decision "
        "is logged but the row never becomes a confirmed mapping."
    ),
}

# Pastel fills for the Confidence column so a reviewer can scan the
# detail sheet visually and tell at a glance which rows still need
# attention. Hex chosen for legibility on the default white background.
CONFIDENCE_FILLS = {
    "HIGH": PatternFill("solid", fgColor="C6EFCE"),       # green
    "MEDIUM": PatternFill("solid", fgColor="FFEB9C"),     # amber
    "LOW": PatternFill("solid", fgColor="FFC7CE"),        # red
    "UNRESOLVED": PatternFill("solid", fgColor="D9D9D9"), # gray
    "EXCLUDED": PatternFill("solid", fgColor="DDEBF7"),   # blue
}

# Same palette keyed by Review Status, so the trust-slice column lights
# up alongside Confidence. ``audited_excluded`` borrows EXCLUDED's blue
# because that's the same audit-only meaning a reviewer is reading.
REVIEW_STATUS_FILLS = {
    "pending_review": CONFIDENCE_FILLS["HIGH"],
    "needs_review": CONFIDENCE_FILLS["MEDIUM"],
    "needs_source": CONFIDENCE_FILLS["UNRESOLVED"],
    "audited_excluded": CONFIDENCE_FILLS["EXCLUDED"],
}

# Per-column widths on the detail sheet. Numbers chosen so the trust
# columns (A–J) and the top-candidate columns (K–U) read without
# horizontal scrolling under typical content; the reviewer-decision
# block (V–Y) stays wide enough to type a free-text note.
_DETAIL_COLUMN_WIDTHS = {
    "A": 12,  # Word ID
    "B": 18,  # Word Location
    "C": 50,  # Word Snippet
    "D": 16,  # Word Raw Token
    "E": 12,  # Word Value
    "F": 10,  # Word Unit
    "G": 12,  # Confidence
    "H": 18,  # Review Status
    "I": 22,  # Placeholder Status
    "J": 18,  # Placeholder
    "K": 16,  # Top Excel Sheet
    "L": 14,  # Top Excel Cell
    "M": 14,  # Top Excel Value
    "N": 30,  # Top Row Context
    "O": 30,  # Top Column Context
    "P": 18,  # Interpretation
    "Q": 12,  # Value Score
    "R": 12,  # Context Score
    "S": 28,  # Overlap Tokens
    "T": 14,  # # Alt Candidates
    "U": 30,  # Note
    "V": 18,  # Reviewer Decision
    "W": 36,  # Reviewer Notes
    "X": 18,  # Confirmed Sheet
    "Y": 14,  # Confirmed Cell
}


def write_mapping_review(
    matches: List[WordMatch],
    out_path: Path,
    word_ids: List[str],
    placeholder_status: Dict[int, str],
) -> Path:
    """Write the ``mapping_review.xlsx`` workbook to ``out_path``.

    ``word_ids`` and ``placeholder_status`` must come from the same
    template-builder run that produced ``auto_mapping.yml``; the validator
    relies on the XLSX agreeing with the YAML on every join key.
    """
    if len(word_ids) != len(matches):
        raise ValueError(
            f"word_ids length {len(word_ids)} != matches length {len(matches)}"
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mapping_review"
    ws.append(REVIEW_HEADERS)
    # Tracked so we can colour Confidence and Review Status cells in a
    # second pass without re-deriving values from the row tuples.
    row_confidences: List[str] = []
    row_review_statuses: List[str] = []
    for i, m in enumerate(matches):
        wn = m.word_number
        top = m.chosen or (m.candidates[0] if m.candidates else None)
        ph_status = placeholder_status.get(i, "skipped_unknown")
        placeholder = "{{ " + word_ids[i] + " }}" if ph_status == "applied" else ""
        review = derive_review_status(m.confidence, ph_status)
        trust_cols = [word_ids[i]]
        trust_tail = [review, ph_status, placeholder]
        # Reviewer-facing columns: blank by default on every learn run.
        # confirm-mapping reads these back; blank must never be promoted
        # to "confirmed" — that contract lives in src.mapping_confirmer.
        reviewer_cols = ["", "", "", ""]
        if top is not None:
            cell = top.cell
            row = trust_cols + [
                wn.location,
                _truncate(wn.snippet, 200),
                wn.raw,
                wn.value,
                wn.unit or "",
                m.confidence,
            ] + trust_tail + [
                cell.sheet,
                cell.address,
                cell.numeric_value,
                " | ".join(cell.row_context[:4]),
                " | ".join(cell.column_context[:4]),
                top.interpretation,
                round(top.value_score, 3),
                round(top.context_score, 3),
                ", ".join(top.overlap_tokens),
                max(0, len(m.candidates) - 1),
                m.note,
            ] + reviewer_cols
        else:
            row = trust_cols + [
                wn.location,
                _truncate(wn.snippet, 200),
                wn.raw,
                wn.value,
                wn.unit or "",
                m.confidence,
            ] + trust_tail + [
                "", "", "", "", "", "", "", "", "", 0, m.note,
            ] + reviewer_cols
        ws.append(row)
        row_confidences.append(m.confidence)
        row_review_statuses.append(review)

    _format_detail_sheet(ws, row_confidences, row_review_statuses)

    # Summary sheet goes at tab position 0 so a reviewer opening the
    # workbook in Excel/LibreOffice sees the aggregate-counts tab on
    # the left, but mapping_review stays the *active* sheet so the
    # programmatic readers (confirm-mapping, validate-artifacts,
    # pilot-summary) — all of which use ``wb.active`` — continue to
    # land on the detail rows without modification.
    summary_ws = wb.create_sheet(SUMMARY_SHEET_TITLE, 0)
    _write_summary_sheet(summary_ws, matches, row_review_statuses)
    wb.active = wb.index(ws)

    wb.save(str(out_path))
    return out_path


def _format_detail_sheet(
    ws,
    row_confidences: List[str],
    row_review_statuses: List[str],
) -> None:
    """Apply reviewer-friendly formatting on the mapping_review sheet.

    None of this changes any cell value — the contract with
    ``validate-artifacts`` and ``confirm-mapping`` is that every column
    and value in :data:`REVIEW_HEADERS` is untouched. We only set
    visual properties (column widths, freeze, autofilter, fills, header
    font) so a reviewer can navigate the sheet without losing the
    header row, the join-key column, or the audit story.
    """
    for col_letter, width in _DETAIL_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Freeze the header row AND the Word ID column. A reviewer paging
    # through hundreds of rows still sees both the column titles and
    # the join key for the row they're inspecting.
    ws.freeze_panes = "B2"

    # Autofilter the whole data range so a reviewer can sort by
    # Confidence or filter to LOW/UNRESOLVED without typing a formula.
    last_col = get_column_letter(len(REVIEW_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    confidence_col = REVIEW_HEADERS.index("Confidence") + 1
    review_col = REVIEW_HEADERS.index("Review Status") + 1
    word_id_col = REVIEW_HEADERS.index("Word ID") + 1
    for row_idx, (conf, review) in enumerate(
        zip(row_confidences, row_review_statuses), start=2
    ):
        fill = CONFIDENCE_FILLS.get(conf)
        if fill is not None:
            # Confidence + Word ID share the fill so the reviewer can
            # tell from the join-key column alone what bucket the row
            # is in without scrolling sideways to find Confidence.
            ws.cell(row=row_idx, column=confidence_col).fill = fill
            ws.cell(row=row_idx, column=word_id_col).fill = fill
        review_fill = REVIEW_STATUS_FILLS.get(review)
        if review_fill is not None:
            ws.cell(row=row_idx, column=review_col).fill = review_fill


def _write_summary_sheet(
    ws,
    matches: List[WordMatch],
    review_statuses: List[str],
) -> None:
    """Render the Summary tab: aggregate counts + next-action guidance.

    The schema is intentionally column-stable (label in A, count in B,
    guidance in C) so a future reviewer-tooling layer can parse it
    without depending on cell formatting. Cells that count things hold
    real integers — not strings — so a spreadsheet user can SUM them.
    """
    total = len(matches)
    confidence_counts: Dict[str, int] = {label: 0 for label in CONFIDENCE_ORDER}
    for m in matches:
        confidence_counts[m.confidence] = confidence_counts.get(m.confidence, 0) + 1
    review_status_counts: Dict[str, int] = {
        "pending_review": 0,
        "needs_review": 0,
        "needs_source": 0,
        "audited_excluded": 0,
    }
    for rs in review_statuses:
        review_status_counts[rs] = review_status_counts.get(rs, 0) + 1
    excluded = confidence_counts.get("EXCLUDED", 0)
    eligible = total - excluded
    mapped = confidence_counts.get("HIGH", 0) + confidence_counts.get("MEDIUM", 0)

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True)

    ws["A1"] = "mapping_review.xlsx — reviewer summary"
    ws["A1"].font = title_font
    ws["A2"] = (
        "Aggregate counts and per-status next actions. "
        "The mapping_review tab has the per-row details; this tab is the "
        "operator-pilot reviewer's entry point."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    row = 4
    ws.cell(row=row, column=1, value="Totals").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Total visible Word numbers")
    ws.cell(row=row, column=2, value=total)
    row += 1
    ws.cell(row=row, column=1, value="Eligible for mapping (excludes EXCLUDED)")
    ws.cell(row=row, column=2, value=eligible)
    row += 1
    ws.cell(row=row, column=1, value="Mapped (HIGH + MEDIUM)")
    ws.cell(row=row, column=2, value=mapped)
    row += 2

    ws.cell(row=row, column=1, value="Counts by Confidence").font = section_font
    ws.cell(row=row, column=2, value="Count").font = section_font
    ws.cell(row=row, column=3, value="Next action").font = section_font
    row += 1
    for label in CONFIDENCE_ORDER:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=confidence_counts.get(label, 0))
        ws.cell(
            row=row, column=3, value=CONFIDENCE_GUIDANCE[label]
        ).alignment = Alignment(wrap_text=True, vertical="top")
        fill = CONFIDENCE_FILLS.get(label)
        if fill is not None:
            ws.cell(row=row, column=1).fill = fill
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="Counts by Review Status").font = section_font
    ws.cell(row=row, column=2, value="Count").font = section_font
    row += 1
    for label in ("pending_review", "needs_review", "needs_source", "audited_excluded"):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=review_status_counts.get(label, 0))
        fill = REVIEW_STATUS_FILLS.get(label)
        if fill is not None:
            ws.cell(row=row, column=1).fill = fill
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 80
    ws.freeze_panes = "A4"


def write_confidence_report(
    matches: List[WordMatch],
    summary: CoverageSummary,
    out_path: Path,
) -> Path:
    """Write a Markdown overview that leads with what needs review."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# Learn-mode confidence report")
    lines.append("")
    lines.append(f"- Total visible Word numbers: **{summary.total}**")
    lines.append(f"  - EXCLUDED by policy: {summary.by_confidence.get('EXCLUDED', 0)}")
    lines.append(f"- Eligible for mapping: **{summary.eligible}**")
    lines.append(f"  - HIGH: {summary.by_confidence.get('HIGH', 0)}")
    lines.append(f"  - MEDIUM: {summary.by_confidence.get('MEDIUM', 0)}")
    lines.append(f"  - LOW: {summary.by_confidence.get('LOW', 0)}")
    lines.append(f"  - UNRESOLVED: {summary.by_confidence.get('UNRESOLVED', 0)}")
    lines.append(f"- Mapped (HIGH+MEDIUM): **{summary.mapped} ({summary.coverage_ratio:.1%} of eligible)**")
    lines.append("")
    lines.append("> This report exists so a human can see at a glance which Word numbers")
    lines.append("> are not yet trustworthy. Unresolved and low-confidence entries must")
    lines.append("> be reviewed before any mapping is confirmed for production rendering.")
    lines.append("> EXCLUDED entries were skipped by explicit policy (e.g. date markers");
    lines.append("> like '5月' or '2026年'); the reviewer should still scan them to")
    lines.append("> confirm the policy applied correctly.")
    lines.append("")

    lines.append("## UNRESOLVED")
    if not summary.unresolved:
        # Scope the claim to "eligible" — saying "every Word number has a
        # source" would be a false success when EXCLUDED rows exist, since
        # those were never matched and have no candidates.
        msg = "_None — every eligible Word number has at least one candidate Excel source._"
        excluded_count = summary.by_confidence.get("EXCLUDED", 0)
        if excluded_count:
            msg += f" ({excluded_count} number(s) were EXCLUDED by policy — see section below.)"
        lines.append(msg)
    else:
        for m in summary.unresolved:
            wn = m.word_number
            lines.append(f"- `{wn.location}` — raw `{wn.raw}` (value={wn.value}, unit={wn.unit or '∅'})")
            lines.append(f"    - snippet: {_truncate(wn.snippet, 160)}")
    lines.append("")

    lines.append("## LOW confidence")
    if not summary.low_confidence:
        lines.append("_None._")
    else:
        for m in summary.low_confidence:
            wn = m.word_number
            top = m.chosen or (m.candidates[0] if m.candidates else None)
            lines.append(f"- `{wn.location}` — raw `{wn.raw}`")
            lines.append(f"    - snippet: {_truncate(wn.snippet, 160)}")
            if top is not None:
                lines.append(f"    - top guess: {top.cell.sheet}!{top.cell.address} = {top.cell.numeric_value} ({top.interpretation})")
    lines.append("")

    lines.append("## Ambiguous picks (MEDIUM with ties)")
    if not summary.ambiguous:
        lines.append("_None._")
    else:
        for m in summary.ambiguous:
            wn = m.word_number
            lines.append(f"- `{wn.location}` — raw `{wn.raw}`: {m.note}")
            for c in m.candidates[:3]:
                lines.append(f"    - {c.cell.sheet}!{c.cell.address} = {c.cell.numeric_value} (row_ctx={c.cell.row_context[:2]}, col_ctx={c.cell.column_context[:2]})")
    lines.append("")

    lines.append("## EXCLUDED by explicit policy")
    if not summary.excluded:
        lines.append("_None._")
    else:
        lines.append(f"_{len(summary.excluded)} number(s) captured for audit but skipped by the matcher._")
        lines.append("")
        for m in summary.excluded:
            wn = m.word_number
            lines.append(f"- `{wn.location}` — raw `{wn.raw}` — {m.note}")
            lines.append(f"    - snippet: {_truncate(wn.snippet, 160)}")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def _truncate(text: str, limit: int) -> str:
    text = (text or "").replace("\n", " ").strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"
