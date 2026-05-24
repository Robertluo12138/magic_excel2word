"""Run-mode preview: deterministically extract values from a NEW Excel
workbook using a previously human-confirmed mapping, and emit a run
validation artifact for review.

This is NOT a Word renderer. It validates that the confirmed mappings
still resolve cleanly against the new period's Excel file before any
production Word output is attempted. The artifact shows one row per
confirmed ``word_id`` with the raw Excel value, the post-transform
generated value, and the per-row status — exactly the substrate a
reviewer needs to spot drift between learn-time and run-time.

Fail-loud contracts (the CLI exits non-zero unless every condition holds):

* ``confirmed_mapping.yml`` must be complete (``summary.complete: true``,
  zero ``review_required`` entries, ``summary.allow_incomplete: false``,
  and at least one confirmed mapping). A partial or exploratory file may
  be useful at learn time; the run-preview gate refuses to use one.
* Every confirmed entry must name a ``confirmed_source`` ``(sheet, cell)``
  that exists in the new Excel workbook.
* Every named cell must hold a numeric value (booleans and strings fail).
* The transform interpretation recorded on each entry must be one the
  v1 transform table knows how to invert. Unknown interpretations fail
  rather than silently producing a wrong rendered number.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set

import openpyxl
import yaml

# Inverse of number_normalizer.candidate_values for run-mode rendering.
# At learn time the matcher chose interpretation X to bridge the Word
# ``word_raw`` value and the Excel cell value via the relation
# ``excel_value == word_raw * FACTOR``. To regenerate the value the Word
# document would display from the new Excel cell, invert that:
# ``generated == excel_value / FACTOR``.
#
# Keep this table in sync with ``number_normalizer.candidate_values`` —
# any new interpretation added there must also be invertible here, or
# it must be excluded from confirmable rows upstream.
_INTERPRETATION_FACTORS: Dict[str, float] = {
    "as_written": 1.0,
    "万元→base_unit": 10_000.0,
    "万人→base_unit": 10_000.0,
    "万次→base_unit": 10_000.0,
    "万个→base_unit": 10_000.0,
    "万单→base_unit": 10_000.0,
    "万→base_unit": 10_000.0,
    "亿元→base_unit": 100_000_000.0,
    "亿单→base_unit": 100_000_000.0,
    "千元→元": 1_000.0,
    "百万元→元": 1_000_000.0,
    "%→decimal": 1.0 / 100.0,
    "‰→decimal": 1.0 / 1_000.0,
    "base→万": 1.0 / 10_000.0,
}


# Per-row run statuses. ``ok`` is the only success state; everything
# else fails the gate and the command exits non-zero. The status is
# surfaced in the artifact so a reviewer can see exactly which rows
# broke without re-running the pipeline.
STATUS_OK = "ok"
STATUS_MISSING_SOURCE = "missing_confirmed_source"
STATUS_MISSING_SHEET = "missing_sheet"
STATUS_MISSING_CELL = "missing_cell"
STATUS_NON_NUMERIC = "non_numeric_cell"
STATUS_MISSING_TRANSFORM = "missing_transform"
STATUS_TRANSFORM_UNKNOWN = "transform_unknown"


@dataclass
class PreviewRow:
    """One resolved (or failed-to-resolve) confirmed mapping entry."""
    word_id: str
    location: str
    word_snippet: str
    word_raw: str
    word_unit: str
    source_sheet: str
    source_cell: str
    raw_excel_value: Optional[float]
    generated_value: Optional[float]
    transform_interpretation: str
    confidence: str
    status: str
    detail: str = ""


@dataclass
class PreviewReport:
    rows: List[PreviewRow] = field(default_factory=list)
    fatal_errors: List[str] = field(default_factory=list)

    @property
    def failures(self) -> List[PreviewRow]:
        return [r for r in self.rows if r.status != STATUS_OK]

    @property
    def ok(self) -> bool:
        return not self.fatal_errors and not self.failures and bool(self.rows)


def run_preview(excel_path: Path, confirmed_path: Path) -> PreviewReport:
    """Read the confirmed mapping + the new Excel and resolve every row.

    Returns a :class:`PreviewReport`. The CLI decides how to render exit
    codes; this function just classifies. Pre-condition: both files
    exist (the CLI guards that for friendlier error messages).
    """
    report = PreviewReport()

    try:
        confirmed_doc = yaml.safe_load(confirmed_path.read_text(encoding="utf-8"))
    except Exception as exc:
        report.fatal_errors.append(f"cannot parse {confirmed_path.name}: {exc}")
        return report
    if not isinstance(confirmed_doc, dict):
        report.fatal_errors.append(
            f"{confirmed_path.name} is not a YAML mapping"
        )
        return report

    # Completeness gate. Fail-CLOSED: refuse unless the file proves it is
    # complete. A confirmed_mapping.yml that is missing the summary
    # block, missing the ``complete`` flag, has ``complete`` set to
    # anything other than the boolean ``True``, has any ``review_required``
    # row, or carries ``allow_incomplete: true`` is exploratory by
    # definition. Running production-like extraction against it would
    # silently elide whatever the reviewer has not yet signed off on.
    #
    # The strict ``is True`` / ``is not True`` checks matter: a typo or
    # missing field would produce ``None``, which ``is False`` would
    # silently let through. ``is not True`` rejects ``None``, ``"true"``,
    # ``1``, and every other near-miss.
    summary = confirmed_doc.get("summary")
    if not isinstance(summary, dict):
        report.fatal_errors.append(
            f"{confirmed_path.name} is missing the top-level 'summary' "
            "block — refuse to run preview against a file whose "
            "completeness cannot be verified."
        )
        return report

    review_required = confirmed_doc.get("review_required") or []
    if review_required:
        report.fatal_errors.append(
            f"{confirmed_path.name} has {len(review_required)} review_required "
            "entries — refuse to run preview against a partial mapping. "
            "Re-run confirm-mapping after the reviewer fills in the blocking "
            "rows."
        )
        return report
    # allow_incomplete is checked before complete so the error message
    # names the explicit user choice rather than the derived flag — the
    # reviewer needs to know they opted into a partial file.
    if summary.get("allow_incomplete") is True:
        report.fatal_errors.append(
            f"{confirmed_path.name} summary.allow_incomplete is true — the "
            "file was written via the --allow-incomplete escape hatch; "
            "refuse to run preview against an exploratory partial mapping."
        )
        return report
    if summary.get("complete") is not True:
        report.fatal_errors.append(
            f"{confirmed_path.name} summary.complete is "
            f"{summary.get('complete')!r}, must be the boolean true — "
            "refuse to run preview against a mapping that cannot prove "
            "it is complete."
        )
        return report

    confirmed = confirmed_doc.get("confirmed_mappings")
    if not isinstance(confirmed, list):
        report.fatal_errors.append(
            f"{confirmed_path.name} missing top-level 'confirmed_mappings' list"
        )
        return report
    if not confirmed:
        report.fatal_errors.append(
            f"{confirmed_path.name} has no confirmed mappings — refuse to "
            "generate a preview from an empty file."
        )
        return report

    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    except Exception as exc:
        report.fatal_errors.append(f"cannot open {excel_path.name}: {exc}")
        return report

    try:
        sheet_names: Set[str] = set(wb.sheetnames)
        # No-silent-omission invariant: one PreviewRow per confirmed
        # entry, in input order. Even a row that fails resolution gets
        # surfaced so the reviewer sees which word_id broke.
        for entry in confirmed:
            report.rows.append(_resolve_entry(entry, wb, sheet_names))
    finally:
        wb.close()

    return report


def _resolve_entry(entry: Dict, wb, sheet_names: Set[str]) -> PreviewRow:
    word_id = str(entry.get("word_id") or "")
    location = str(entry.get("location") or "")
    context = entry.get("context") or {}
    word_snippet = str(context.get("snippet") or "") if isinstance(context, dict) else ""
    word_raw = str(entry.get("raw") or "")
    word_unit = str(entry.get("unit") or "")
    confidence = str(entry.get("confidence") or entry.get("status") or "")

    source = entry.get("confirmed_source") or {}
    if not isinstance(source, dict):
        source = {}
    source_sheet = str(source.get("sheet") or "")
    source_cell = str(source.get("address") or "")
    transform = entry.get("transform") if isinstance(entry.get("transform"), dict) else {}
    interpretation = str((transform or {}).get("interpretation") or "")

    base = PreviewRow(
        word_id=word_id,
        location=location,
        word_snippet=word_snippet,
        word_raw=word_raw,
        word_unit=word_unit,
        source_sheet=source_sheet,
        source_cell=source_cell,
        raw_excel_value=None,
        generated_value=None,
        transform_interpretation=interpretation,
        confidence=confidence,
        status=STATUS_OK,
    )

    if not source_sheet or not source_cell:
        base.status = STATUS_MISSING_SOURCE
        base.detail = "confirmed_source must carry both sheet and address"
        return base

    if source_sheet not in sheet_names:
        base.status = STATUS_MISSING_SHEET
        base.detail = f"sheet {source_sheet!r} not found in new workbook"
        return base

    ws = wb[source_sheet]
    try:
        cell = ws[source_cell]
    except Exception as exc:
        # Malformed cell address (e.g. "ZZZ" without a row, or weird
        # characters). Surface as missing_cell so it's grouped with the
        # other "can't read this cell" failures rather than crashing.
        base.status = STATUS_MISSING_CELL
        base.detail = f"cannot read cell {source_cell!r}: {exc}"
        return base

    raw = getattr(cell, "value", None)
    if raw is None:
        base.status = STATUS_MISSING_CELL
        base.detail = f"cell {source_sheet}!{source_cell} is empty"
        return base

    # bool is a subclass of int — reject explicitly so TRUE/FALSE cells
    # don't masquerade as numeric. Matches excel_profiler.profile_workbook.
    if isinstance(raw, bool) or not isinstance(raw, (int, float)):
        base.status = STATUS_NON_NUMERIC
        base.detail = (
            f"cell {source_sheet}!{source_cell} holds non-numeric value "
            f"{raw!r} (type={type(raw).__name__})"
        )
        return base

    excel_value = float(raw)
    base.raw_excel_value = excel_value

    if not interpretation:
        base.status = STATUS_MISSING_TRANSFORM
        base.detail = "confirmed entry has no transform.interpretation"
        return base
    if interpretation not in _INTERPRETATION_FACTORS:
        base.status = STATUS_TRANSFORM_UNKNOWN
        base.detail = (
            f"transform interpretation {interpretation!r} not in the v1 "
            "table; update _INTERPRETATION_FACTORS or fix the confirmed "
            "mapping upstream"
        )
        return base

    factor = _INTERPRETATION_FACTORS[interpretation]
    # generated value is the unrounded numeric value; a future renderer
    # is responsible for matching the report's display formatting
    # (e.g. ``23456.789 → "23,456.79万元"``).
    base.generated_value = excel_value / factor
    return base


def write_run_validation(report: PreviewReport, out_dir: Path) -> Path:
    """Write ``run_validation.xlsx`` under ``out_dir`` and return its path.

    Always written when called — even when some rows failed — so the
    reviewer can inspect every confirmed ``word_id`` (no silent omission)
    and see exactly which ones broke. The CLI handles the non-zero exit.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / "run_validation.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "run_validation"
    headers = [
        "Word ID",
        "Word Location",
        "Word Context",
        "Word Raw Token",
        "Word Unit",
        "Source Sheet",
        "Source Cell",
        "Raw Excel Value",
        "Generated Value",
        "Transform Interpretation",
        "Confidence",
        "Status",
        "Detail",
    ]
    ws.append(headers)
    for row in report.rows:
        ws.append([
            row.word_id,
            row.location,
            row.word_snippet,
            row.word_raw,
            row.word_unit,
            row.source_sheet,
            row.source_cell,
            row.raw_excel_value,
            row.generated_value,
            row.transform_interpretation,
            row.confidence,
            row.status,
            row.detail,
        ])
    wb.save(str(path))
    return path


def format_console_summary(report: PreviewReport, out_path: Optional[Path]) -> str:
    """Human-readable summary for the CLI."""
    ok = sum(1 for r in report.rows if r.status == STATUS_OK)
    lines = [
        "run-preview summary",
        "===================",
        f"  total confirmed rows : {len(report.rows)}",
        f"  ok                   : {ok}",
        f"  failed               : {len(report.failures)}",
    ]
    if out_path is not None:
        lines.append(f"  written to           : {out_path}")
    if report.failures:
        lines.append("")
        lines.append("Failures (first 10):")
        for row in report.failures[:10]:
            lines.append(
                f"  - {row.word_id} @ {row.source_sheet}!{row.source_cell} "
                f"[{row.status}] {row.detail}"
            )
        if len(report.failures) > 10:
            lines.append(f"  … and {len(report.failures) - 10} more")
    if report.fatal_errors:
        lines.append("")
        lines.append("FATAL:")
        for err in report.fatal_errors:
            lines.append(f"  - {err}")
    return "\n".join(lines)
