"""Tests for the ``confirm-mapping`` reviewer-handoff gate.

Contracts under test (mirrored from src.mapping_confirmer's docstring):

* Blank ``Reviewer Decision`` is never promoted to confirmed.
* Explicit ``confirm`` on a HIGH/MEDIUM row whose source matches the
  matcher's recommended pick (or one of the alternatives) lands in
  ``confirmed_mappings`` with the full traceability chain preserved.
* LOW and UNRESOLVED rows stay in ``review_required`` regardless of
  decision — a reviewer-side checkbox cannot override the matcher's
  inability to find/score a safe candidate.
* A reviewer override that names a (sheet, cell) the matcher never
  produced is invalid: the row goes to ``review_required`` and is NOT
  silently dropped.
* No silent omission: every YAML mapping ends up in exactly one of the
  three buckets, accounted for in the summary.
* EXCLUDED rows always go to ``audit_only_excluded``; the reviewer's
  decision is recorded for the audit trail but never promotes them.
* CLI exits non-zero when ``review_required`` is non-empty, unless
  ``--allow-incomplete`` is set.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pytest
import yaml

from src.main import main as cli_main
from src.mapping_confirmer import (
    KNOWN_DECISIONS,
    confirm_mappings,
    write_confirmed_yaml,
)
from src.mapping_reviewer import REVIEW_HEADERS, REVIEWER_COLUMNS
from src.synthetic_generator import generate


# ---------------------------------------------------------------------------
# Fixture builders — tiny, hand-built auto_mapping + review xlsx
# ---------------------------------------------------------------------------

def _build_auto_yaml(entries: List[Dict]) -> Dict:
    by_conf: Dict[str, int] = {
        "HIGH": 0, "MEDIUM": 0, "LOW": 0, "UNRESOLVED": 0, "EXCLUDED": 0,
    }
    for e in entries:
        by_conf[e["status"]] = by_conf.get(e["status"], 0) + 1
    return {
        "schema_version": 1,
        "summary": {
            "total": len(entries),
            "by_confidence": by_conf,
            "placeholders_applied": sum(
                1 for e in entries if e.get("placeholder_status") == "applied"
            ),
        },
        "mappings": entries,
    }


def _entry(
    word_id: str,
    *,
    status: str,
    location: str = "paragraph:0",
    raw: str = "100",
    value: float = 100.0,
    unit: str = "元",
    recommended: Optional[Dict] = None,
    alternatives: Optional[List[Dict]] = None,
    placeholder_status: Optional[str] = None,
    note: str = "",
) -> Dict:
    """Build one auto_mapping.yml entry that resembles real output."""
    if placeholder_status is None:
        placeholder_status = (
            "applied" if status in ("HIGH", "MEDIUM") else f"skipped_{status.lower()}"
        )
    review_status = {
        "HIGH": "pending_review",
        "MEDIUM": "pending_review",
        "LOW": "needs_review",
        "UNRESOLVED": "needs_source",
        "EXCLUDED": "audited_excluded",
    }[status]
    placeholder_value = (
        "{{ " + word_id + " }}" if placeholder_status == "applied" else None
    )
    return {
        "word_id": word_id,
        "location": location,
        "raw": raw,
        "value": value,
        "unit": unit,
        "sign": 1,
        "context": {"snippet": f"{raw} 收入", "label_context": []},
        "status": status,
        "confidence": status,
        "review_status": review_status,
        "note": note,
        "placeholder": placeholder_value,
        "placeholder_status": placeholder_status,
        "recommended_source": recommended,
        "transform": (
            {
                "interpretation": "as_written",
                "value_score": 1.0,
                "context_score": 1.0,
                "overlap_tokens": ["收入"],
            }
            if recommended is not None
            else None
        ),
        "alternatives": alternatives or [],
    }


def _source(sheet: str, address: str, value: float) -> Dict:
    return {
        "sheet": sheet,
        "address": address,
        "value": value,
        "row_context": ["收入"],
        "column_context": ["2026年5月"],
    }


def _alt(sheet: str, address: str, value: float) -> Dict:
    return {
        **_source(sheet, address, value),
        "interpretation": "as_written",
        "value_score": 0.85,
        "context_score": 0.45,
        "overlap_tokens": ["万元"],
    }


def _write_auto_yaml(path: Path, entries: List[Dict]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump(
            _build_auto_yaml(entries),
            allow_unicode=True, sort_keys=False, width=1000,
        ),
        encoding="utf-8",
    )
    return path


def _write_review_xlsx(
    path: Path,
    entries: List[Dict],
    decisions: Dict[str, Dict[str, str]],
) -> Path:
    """Write a mapping_review.xlsx with one row per entry.

    ``decisions`` maps ``word_id`` → dict with optional keys
    ``decision``, ``notes``, ``sheet``, ``cell``. Missing word_ids are
    written with blank reviewer columns (the default learn-mode state).
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mapping_review"
    ws.append(REVIEW_HEADERS)
    col_idx = {h: REVIEW_HEADERS.index(h) for h in REVIEW_HEADERS}
    for entry in entries:
        row = [""] * len(REVIEW_HEADERS)
        row[col_idx["Word ID"]] = entry["word_id"]
        row[col_idx["Word Location"]] = entry["location"]
        row[col_idx["Word Raw Token"]] = entry["raw"]
        row[col_idx["Confidence"]] = entry["status"]
        row[col_idx["Review Status"]] = entry["review_status"]
        row[col_idx["Placeholder Status"]] = entry["placeholder_status"]
        row[col_idx["Placeholder"]] = entry["placeholder"] or ""
        rec = entry.get("recommended_source") or {}
        row[col_idx["Top Excel Sheet"]] = rec.get("sheet", "")
        row[col_idx["Top Excel Cell"]] = rec.get("address", "")
        # The four reviewer-facing columns: blank unless decisions says otherwise.
        d = decisions.get(entry["word_id"], {})
        row[col_idx["Reviewer Decision"]] = d.get("decision", "")
        row[col_idx["Reviewer Notes"]] = d.get("notes", "")
        row[col_idx["Confirmed Sheet"]] = d.get("sheet", "")
        row[col_idx["Confirmed Cell"]] = d.get("cell", "")
        ws.append(row)
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# mapping_reviewer XLSX must ship the four reviewer columns blank.
# ---------------------------------------------------------------------------

def test_mapping_review_xlsx_has_reviewer_columns_blank_by_default(tmp_path: Path):
    samples = tmp_path / "samples"
    out = tmp_path / "out"
    generate(samples)
    rc = cli_main([
        "learn",
        "--excel", str(samples / "historical.xlsx"),
        "--word", str(samples / "finished_report.docx"),
        "--out", str(out),
    ])
    assert rc == 0
    wb = openpyxl.load_workbook(str(out / "mapping_review.xlsx"))
    ws = wb.active
    header = [c.value for c in ws[1]]
    # Every reviewer column must be present in the header.
    for col in REVIEWER_COLUMNS:
        assert col in header, f"learn-mode XLSX missing reviewer column {col!r}"
    indices = [header.index(col) + 1 for col in REVIEWER_COLUMNS]
    # And every data row must have those four cells blank — never
    # pre-populated, otherwise blank-as-not-confirmed loses meaning.
    for r in range(2, ws.max_row + 1):
        for c in indices:
            v = ws.cell(row=r, column=c).value
            assert v in (None, ""), (
                f"reviewer column at row {r} col {c} was pre-populated: {v!r}; "
                "learn mode must leave reviewer columns blank by default"
            )


# ---------------------------------------------------------------------------
# Blank decisions must NOT be promoted.
# ---------------------------------------------------------------------------

def test_blank_decision_is_never_confirmed(tmp_path: Path):
    entries = [
        _entry("word_0001", status="HIGH", recommended=_source("月度", "B2", 100.0)),
        _entry("word_0002", status="MEDIUM", recommended=_source("月度", "B3", 200.0)),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(tmp_path / "review.xlsx", entries, decisions={})
    report = confirm_mappings(auto, review)

    assert report.confirmed == []
    assert {r["word_id"] for r in report.review_required} == {"word_0001", "word_0002"}
    assert all(r["reason"] == "blank_decision" for r in report.review_required)


def test_cli_blank_decisions_fail_without_allow_incomplete(tmp_path: Path, capsys):
    entries = [
        _entry("word_0001", status="HIGH", recommended=_source("月度", "B2", 100.0)),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(tmp_path / "review.xlsx", entries, decisions={})
    out = tmp_path / "confirmed.yml"

    rc = cli_main([
        "confirm-mapping",
        "--auto", str(auto),
        "--review", str(review),
        "--out", str(out),
    ])

    assert rc == 5, "blank decision must fail the default gate"
    assert out.exists(), "confirmed_mapping.yml is still written so reviewer can see what's missing"
    data = yaml.safe_load(out.read_text(encoding="utf-8"))
    assert data["summary"]["confirmed"] == 0
    assert data["summary"]["review_required"] == 1
    assert data["summary"]["complete"] is False
    captured = capsys.readouterr()
    assert "review_required" in captured.out
    assert "blank_decision" in captured.out


def test_cli_allow_incomplete_returns_zero_with_review_required(tmp_path: Path):
    entries = [
        _entry("word_0001", status="HIGH", recommended=_source("月度", "B2", 100.0)),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(tmp_path / "review.xlsx", entries, decisions={})
    out = tmp_path / "confirmed.yml"

    rc = cli_main([
        "confirm-mapping",
        "--auto", str(auto),
        "--review", str(review),
        "--out", str(out),
        "--allow-incomplete",
    ])
    assert rc == 0
    data = yaml.safe_load(out.read_text(encoding="utf-8"))
    assert data["summary"]["allow_incomplete"] is True
    assert data["summary"]["review_required"] == 1


# ---------------------------------------------------------------------------
# Confirmed rows succeed and preserve traceability.
# ---------------------------------------------------------------------------

def test_confirm_decision_with_recommended_source_succeeds(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="HIGH", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "confirm", "notes": "looks right"}},
    )
    report = confirm_mappings(auto, review)

    assert report.review_required == []
    [conf] = report.confirmed
    assert conf["word_id"] == "word_0001"
    assert conf["location"] == "paragraph:0"
    assert conf["confirmed_source"] == {"sheet": "月度", "address": "B2", "value": 100.0}
    assert conf["source_origin"] == "recommended"
    assert conf["reviewer_decision"] == "confirm"
    assert conf["reviewer_notes"] == "looks right"
    # Full traceability: word_id, location, transform, recommended_source,
    # confirmed_source, reviewer decision/notes, confidence, status.
    for required in (
        "word_id", "location", "raw", "value", "unit", "status",
        "confidence", "review_status", "transform", "recommended_source",
        "confirmed_source", "source_origin", "reviewer_decision",
        "reviewer_notes", "placeholder", "placeholder_status",
    ):
        assert required in conf, f"confirmed entry missing field {required!r}"


def test_reviewer_override_matching_alternative_succeeds(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    alt = _alt("周度", "B6", 100.0)
    entries = [
        _entry("word_0001", status="HIGH", recommended=rec, alternatives=[alt]),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {
            "decision": "confirm", "sheet": "周度", "cell": "B6",
        }},
    )
    report = confirm_mappings(auto, review)
    [conf] = report.confirmed
    assert conf["confirmed_source"] == {"sheet": "周度", "address": "B6", "value": 100.0}
    assert conf["source_origin"] == "reviewer_override:alternative"


def test_reviewer_override_matching_recommended_succeeds(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="HIGH", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {
            "decision": "confirm", "sheet": "月度", "cell": "B2",
        }},
    )
    report = confirm_mappings(auto, review)
    [conf] = report.confirmed
    assert conf["source_origin"] == "reviewer_override:recommended"


# ---------------------------------------------------------------------------
# Invalid override cells must fail loudly.
# ---------------------------------------------------------------------------

def test_invalid_override_cell_fails(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    alt = _alt("周度", "B6", 100.0)
    entries = [
        _entry("word_0001", status="HIGH", recommended=rec, alternatives=[alt]),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {
            "decision": "confirm", "sheet": "渠道", "cell": "Z99",
        }},
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    [req] = report.review_required
    assert req["word_id"] == "word_0001"
    assert req["reason"].startswith("invalid_override")
    assert req["reviewer_confirmed_sheet"] == "渠道"
    assert req["reviewer_confirmed_cell"] == "Z99"


def test_high_row_with_skipped_placeholder_cannot_be_confirmed(tmp_path: Path):
    """A HIGH/MEDIUM row whose template placeholder was skipped (offset
    drift, raw mismatch) is *non-renderable*: the converted .docx has no
    ``{{ word_NNNN }}`` for a future renderer to substitute into. If
    confirm-mapping promotes such a row, the YAML would claim the
    metric is rendered while the .docx silently omits it — exactly the
    risk CLAUDE.md flags. The gate must keep it in review_required."""
    rec = _source("月度", "B2", 100.0)
    entries = [
        _entry(
            "word_0001",
            status="HIGH",
            recommended=rec,
            placeholder_status="skipped_raw_mismatch",
        ),
        _entry(
            "word_0002",
            status="MEDIUM",
            recommended=rec,
            placeholder_status="skipped_offset_out_of_range",
        ),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={
            "word_0001": {"decision": "confirm"},
            "word_0002": {"decision": "confirm"},
        },
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == [], (
        "non-renderable rows must NOT land in confirmed_mappings even when "
        "the reviewer explicitly marked confirm; "
        f"got: {[c['word_id'] for c in report.confirmed]}"
    )
    reasons = {r["word_id"]: r["reason"] for r in report.review_required}
    assert reasons["word_0001"] == "non_renderable_template_skip:skipped_raw_mismatch"
    assert reasons["word_0002"] == "non_renderable_template_skip:skipped_offset_out_of_range"


def test_partial_override_only_sheet_or_only_cell_is_invalid(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [
        _entry("word_0001", status="HIGH", recommended=rec),
        _entry("word_0002", status="HIGH", recommended=rec),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={
            "word_0001": {"decision": "confirm", "sheet": "月度", "cell": ""},
            "word_0002": {"decision": "confirm", "sheet": "", "cell": "B2"},
        },
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    reasons = {r["word_id"]: r["reason"] for r in report.review_required}
    assert reasons["word_0001"].startswith("incomplete_override")
    assert reasons["word_0002"].startswith("incomplete_override")


# ---------------------------------------------------------------------------
# LOW / UNRESOLVED / EXCLUDED rules.
# ---------------------------------------------------------------------------

def test_unresolved_rows_stay_visible_regardless_of_decision(tmp_path: Path):
    entries = [
        _entry("word_0001", status="UNRESOLVED", recommended=None),
        _entry("word_0002", status="UNRESOLVED", recommended=None),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    # Even a reviewer marking ``confirm`` on UNRESOLVED is not enough —
    # the matcher had no candidate to validate against, so the row stays
    # in review_required with a specific reason.
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={
            "word_0001": {"decision": "confirm"},
            "word_0002": {"decision": ""},
        },
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    assert {r["word_id"] for r in report.review_required} == {"word_0001", "word_0002"}
    reasons = {r["word_id"]: r["reason"] for r in report.review_required}
    assert reasons["word_0001"] == "unresolved_no_candidate"
    assert reasons["word_0002"] == "unresolved_no_candidate"


def test_low_confidence_cannot_be_confirmed_via_checkbox(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="LOW", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "confirm"}},
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    [req] = report.review_required
    assert req["reason"] == "low_confidence_cannot_confirm"


def test_excluded_rows_become_audit_only_even_when_confirmed(tmp_path: Path):
    entries = [
        _entry("word_0001", status="EXCLUDED", recommended=None, raw="2026"),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "confirm", "notes": "year"}},
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    assert report.review_required == []
    [audit] = report.audit_only_excluded
    assert audit["word_id"] == "word_0001"
    assert audit["status"] == "EXCLUDED"
    # Decision is captured for the audit trail even though excluded rows
    # never become confirmed.
    assert audit["reviewer_decision"] == "confirm"
    assert audit["reviewer_notes"] == "year"


# ---------------------------------------------------------------------------
# Reject and unknown decisions land in review_required.
# ---------------------------------------------------------------------------

def test_reject_decision_lands_in_review_required(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="HIGH", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "reject", "notes": "wrong column"}},
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    [req] = report.review_required
    assert req["reason"] == "rejected"
    assert req["reviewer_decision"] == "reject"


def test_unknown_decision_is_invalid_not_silently_confirmed(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="HIGH", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "maybe-later"}},
    )
    report = confirm_mappings(auto, review)
    assert report.confirmed == []
    [req] = report.review_required
    assert req["reason"].startswith("invalid_decision")
    # Sanity: the contract pins exactly which decisions are accepted.
    assert KNOWN_DECISIONS == frozenset({"confirm", "reject", ""})


# ---------------------------------------------------------------------------
# No-silent-omission invariant.
# ---------------------------------------------------------------------------

def test_no_silent_omission_of_eligible_word_numbers(tmp_path: Path):
    """Every YAML mapping must end up in exactly one bucket. This is the
    invariant CLAUDE.md insists on: a confirmed_mapping.yml that drops an
    eligible Word number is the highest-severity bug we can ship."""
    entries = [
        _entry("word_0001", status="HIGH", recommended=_source("月度", "B2", 100.0)),
        _entry("word_0002", status="MEDIUM", recommended=_source("月度", "B3", 200.0)),
        _entry("word_0003", status="LOW", recommended=_source("月度", "B4", 300.0)),
        _entry("word_0004", status="UNRESOLVED", recommended=None),
        _entry("word_0005", status="EXCLUDED", recommended=None, raw="2026"),
    ]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "confirm"}},  # rest blank
    )
    report = confirm_mappings(auto, review)
    assert report.total_accounted == len(entries) == 5
    all_ids = {e["word_id"] for e in entries}
    seen = (
        {c["word_id"] for c in report.confirmed}
        | {r["word_id"] for r in report.review_required}
        | {a["word_id"] for a in report.audit_only_excluded}
    )
    assert seen == all_ids, "every YAML mapping must land in exactly one bucket"


def test_full_corpus_round_trip_with_synthetic_learn(tmp_path: Path):
    """End-to-end: run learn, then confirm-mapping with no decisions —
    every HIGH/MEDIUM must be review_required, every EXCLUDED audit_only,
    UNRESOLVED in review_required, and no row missing."""
    samples = tmp_path / "samples"
    learn_out = tmp_path / "out"
    generate(samples)
    rc = cli_main([
        "learn",
        "--excel", str(samples / "historical.xlsx"),
        "--word", str(samples / "finished_report.docx"),
        "--out", str(learn_out),
    ])
    assert rc == 0

    auto = learn_out / "auto_mapping.yml"
    review = learn_out / "mapping_review.xlsx"
    confirmed = learn_out / "confirmed_mapping.yml"

    rc = cli_main([
        "confirm-mapping",
        "--auto", str(auto),
        "--review", str(review),
        "--out", str(confirmed),
        "--allow-incomplete",
    ])
    assert rc == 0
    data = yaml.safe_load(confirmed.read_text(encoding="utf-8"))
    auto_total = yaml.safe_load(auto.read_text(encoding="utf-8"))["summary"]["total"]
    total = (
        data["summary"]["confirmed"]
        + data["summary"]["review_required"]
        + data["summary"]["audit_only_excluded"]
    )
    assert total == auto_total, "buckets must account for every Word number"
    assert data["summary"]["confirmed"] == 0, "no decisions → no confirmations"
    assert data["summary"]["audit_only_excluded"] > 0, (
        "synthetic corpus has EXCLUDED date markers — they must be audit-only"
    )


def test_cli_confirm_mapping_reviewed_copy_succeeds(tmp_path: Path):
    """Simulate the reviewer copying the XLSX, marking every HIGH/MEDIUM as
    ``confirm``, and re-running. The gate must pass with no
    --allow-incomplete and exit 0."""
    samples = tmp_path / "samples"
    learn_out = tmp_path / "out"
    generate(samples)
    assert cli_main([
        "learn",
        "--excel", str(samples / "historical.xlsx"),
        "--word", str(samples / "finished_report.docx"),
        "--out", str(learn_out),
    ]) == 0

    # Open the XLSX, mark every HIGH/MEDIUM row as ``confirm``, save back.
    review_path = learn_out / "mapping_review.xlsx"
    wb = openpyxl.load_workbook(str(review_path))
    ws = wb.active
    header = [c.value for c in ws[1]]
    conf_col = header.index("Confidence") + 1
    decision_col = header.index("Reviewer Decision") + 1
    eligible_unresolved = 0
    for r in range(2, ws.max_row + 1):
        confidence = ws.cell(row=r, column=conf_col).value
        if confidence in ("HIGH", "MEDIUM"):
            ws.cell(row=r, column=decision_col).value = "confirm"
        elif confidence in ("LOW", "UNRESOLVED"):
            eligible_unresolved += 1
    wb.save(str(review_path))

    confirmed = learn_out / "confirmed_mapping.yml"
    rc = cli_main([
        "confirm-mapping",
        "--auto", str(learn_out / "auto_mapping.yml"),
        "--review", str(review_path),
        "--out", str(confirmed),
    ])
    # The synthetic corpus has 2 UNRESOLVED rows (deliberate). Confirming
    # HIGH/MEDIUM still leaves those in review_required, so the gate
    # must still fail — that's the contract being protected.
    if eligible_unresolved > 0:
        assert rc == 5, (
            "UNRESOLVED rows must keep the gate closed even when all "
            "HIGH/MEDIUM are confirmed"
        )
        data = yaml.safe_load(confirmed.read_text(encoding="utf-8"))
        assert data["summary"]["confirmed"] >= 5
        assert data["summary"]["review_required"] == eligible_unresolved
    else:
        assert rc == 0


# ---------------------------------------------------------------------------
# Fatal-error guards.
# ---------------------------------------------------------------------------

def test_xlsx_word_id_not_in_yaml_is_fatal(tmp_path: Path):
    entries = [_entry("word_0001", status="HIGH", recommended=_source("月度", "B2", 100.0))]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    # Build a review xlsx that includes an unknown word_id.
    bogus = [
        *entries,
        _entry("word_9999", status="HIGH", recommended=_source("月度", "B2", 100.0)),
    ]
    review = _write_review_xlsx(tmp_path / "review.xlsx", bogus, decisions={})
    report = confirm_mappings(auto, review)
    assert report.fatal_errors, "extra XLSX rows must surface as fatal"
    assert any("word_9999" in err for err in report.fatal_errors)


def test_cli_missing_inputs_return_exit_2(tmp_path: Path, capsys):
    rc = cli_main([
        "confirm-mapping",
        "--auto", str(tmp_path / "nope.yml"),
        "--review", str(tmp_path / "nope.xlsx"),
        "--out", str(tmp_path / "confirmed.yml"),
    ])
    assert rc == 2
    err = capsys.readouterr().err
    assert "not found" in err


def test_confirmed_yaml_carries_source_artifact_paths(tmp_path: Path):
    rec = _source("月度", "B2", 100.0)
    entries = [_entry("word_0001", status="HIGH", recommended=rec)]
    auto = _write_auto_yaml(tmp_path / "auto.yml", entries)
    review = _write_review_xlsx(
        tmp_path / "review.xlsx", entries,
        decisions={"word_0001": {"decision": "confirm"}},
    )
    out = tmp_path / "confirmed.yml"
    report = confirm_mappings(auto, review)
    write_confirmed_yaml(
        report, auto, review, out,
        allow_incomplete=False, total_word_numbers=len(entries),
    )
    data = yaml.safe_load(out.read_text(encoding="utf-8"))
    assert data["source_artifacts"]["auto_mapping"] == str(auto)
    assert data["source_artifacts"]["mapping_review"] == str(review)
    assert data["summary"]["total_word_numbers"] == 1
    assert data["summary"]["complete"] is True
