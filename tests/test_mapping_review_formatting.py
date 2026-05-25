"""Tests for the reviewer-friendly ``mapping_review.xlsx`` formatting.

The matcher's per-row contract is pinned elsewhere (``test_learn_smoke``,
``test_validate_artifacts``, ``test_confirm_mapping``). This module is
narrow: it proves the operator-pilot reviewer-experience surfaces stay
in place across a real synthetic learn run — the Summary sheet exists
with the right counts and next-action guidance, the detail sheet keeps
every original column and value, and the workbook stays compatible with
the downstream tools that read ``wb.active``.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from src.main import main as cli_main
from src.mapping_reviewer import (
    CONFIDENCE_FILLS,
    CONFIDENCE_GUIDANCE,
    CONFIDENCE_ORDER,
    REVIEW_HEADERS,
    REVIEW_STATUS_FILLS,
    SUMMARY_SHEET_TITLE,
)
from src.synthetic_generator import generate


def _run_learn(tmp_path: Path) -> Path:
    samples = tmp_path / "samples"
    out = tmp_path / "out"
    generate(samples)
    rc = cli_main([
        "learn",
        "--excel", str(samples / "historical.xlsx"),
        "--word", str(samples / "finished_report.docx"),
        "--out", str(out),
    ])
    assert rc == 0
    return out / "mapping_review.xlsx"


# ---------------------------------------------------------------------------
# Backward-compat: existing columns and downstream `wb.active` contract
# ---------------------------------------------------------------------------

def test_review_sheet_keeps_every_original_header_in_order(tmp_path: Path):
    """The matcher/validator contract is that ``REVIEW_HEADERS`` stays
    exactly the same on disk. A renamed or reordered column breaks
    ``confirm-mapping`` (which looks up by header) and
    ``validate-artifacts`` (which pins position semantics)."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    header_row = [c.value for c in ws[1]]
    assert header_row == REVIEW_HEADERS, (
        "mapping_review header row drifted from REVIEW_HEADERS — this "
        "breaks confirm-mapping's header-keyed reader and "
        "validate-artifacts' positional checks."
    )


def test_mapping_review_is_still_the_active_sheet(tmp_path: Path):
    """``confirm-mapping``, ``validate-artifacts``, and ``pilot-summary``
    all open the workbook with ``ws = wb.active``. Adding the Summary
    sheet at tab position 0 must not move the active sheet, or those
    downstream tools would read the Summary tab instead of the per-row
    review data — silently breaking the human-handoff workflow."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    assert wb.active.title == "mapping_review", (
        f"wb.active is {wb.active.title!r}; confirm-mapping and "
        "validate-artifacts depend on it being the detail sheet."
    )


def test_review_rows_count_and_word_ids_unchanged(tmp_path: Path):
    """Every data row from the legacy single-sheet layout must still be
    present, with the join key (Word ID) intact."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    # Header + at least one data row; sanity-bounded so a regression
    # that empties the sheet fails here loudly instead of silently.
    assert ws.max_row >= 2
    # Every data row's Word ID column must follow the word_NNNN shape.
    import re
    pattern = re.compile(r"^word_\d{4,}$")
    word_ids = [
        ws.cell(row=r, column=1).value
        for r in range(2, ws.max_row + 1)
    ]
    assert all(isinstance(w, str) and pattern.match(w) for w in word_ids), (
        "Word ID column contains non-canonical values; "
        f"first offender: {next((w for w in word_ids if not (isinstance(w, str) and pattern.match(w))), None)!r}"
    )
    assert len(set(word_ids)) == len(word_ids), "Word ID column has duplicates"


# ---------------------------------------------------------------------------
# New reviewer-experience surfaces: Summary sheet
# ---------------------------------------------------------------------------

def test_summary_sheet_exists_at_tab_position_zero(tmp_path: Path):
    """The Summary tab is the operator-pilot reviewer's entry point;
    it must be the leftmost tab so a reviewer who opens the workbook
    discovers it without hunting."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    assert wb.sheetnames[0] == SUMMARY_SHEET_TITLE, (
        f"Summary sheet must be at tab position 0; got: {wb.sheetnames}"
    )
    assert SUMMARY_SHEET_TITLE in wb.sheetnames


def test_summary_sheet_aggregate_counts_match_detail_rows(tmp_path: Path):
    """Counts on the Summary tab are derived from the same matches as
    the detail rows. If they drift, a reviewer reading the Summary
    would form a false impression of how much work remains."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    summary = wb[SUMMARY_SHEET_TITLE]
    detail = wb["mapping_review"]

    # Re-count the detail sheet directly.
    conf_col = REVIEW_HEADERS.index("Confidence") + 1
    review_col = REVIEW_HEADERS.index("Review Status") + 1
    expected_conf: dict = {}
    expected_review: dict = {}
    for r in range(2, detail.max_row + 1):
        c = detail.cell(row=r, column=conf_col).value
        rs = detail.cell(row=r, column=review_col).value
        expected_conf[c] = expected_conf.get(c, 0) + 1
        expected_review[rs] = expected_review.get(rs, 0) + 1
    total = detail.max_row - 1

    # Locate each "Counts by …" section by its header label and read
    # the immediately-following rows. The schema is "label in column A,
    # count in column B" — tests pin that without depending on a
    # specific row offset, so a future spacing tweak doesn't false-fail.
    label_to_row = {
        summary.cell(row=r, column=1).value: r
        for r in range(1, summary.max_row + 1)
    }

    # Totals row.
    assert "Total visible Word numbers" in label_to_row
    assert summary.cell(
        row=label_to_row["Total visible Word numbers"], column=2,
    ).value == total

    # Confidence section.
    for label in CONFIDENCE_ORDER:
        assert label in label_to_row, f"Summary missing confidence row: {label}"
        cell = summary.cell(row=label_to_row[label], column=2).value
        assert cell == expected_conf.get(label, 0), (
            f"Summary Confidence count for {label} disagrees with detail "
            f"sheet: summary={cell}, detail-derived={expected_conf.get(label, 0)}"
        )

    # Review-status section.
    for label in ("pending_review", "needs_review", "needs_source", "audited_excluded"):
        assert label in label_to_row, f"Summary missing review-status row: {label}"
        cell = summary.cell(row=label_to_row[label], column=2).value
        assert cell == expected_review.get(label, 0), (
            f"Summary Review Status count for {label} disagrees with "
            f"detail: summary={cell}, detail-derived={expected_review.get(label, 0)}"
        )


def test_summary_sheet_lists_next_action_guidance_for_every_confidence(tmp_path: Path):
    """A reviewer must see what to do next for each confidence label
    without alt-tabbing to the README. Forgetting to surface the
    guidance is the regression this test catches."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    summary = wb[SUMMARY_SHEET_TITLE]
    guidance_by_label: dict = {}
    for r in range(1, summary.max_row + 1):
        label = summary.cell(row=r, column=1).value
        if label in CONFIDENCE_ORDER:
            guidance_by_label[label] = summary.cell(row=r, column=3).value
    for label in CONFIDENCE_ORDER:
        text = guidance_by_label.get(label) or ""
        # The full guidance string is the contract; cross-check against
        # the source-of-truth constant so a future copy edit in one
        # place can't silently desync the other.
        assert text == CONFIDENCE_GUIDANCE[label], (
            f"Summary next-action guidance for {label} disagrees with "
            f"CONFIDENCE_GUIDANCE: {text!r}"
        )


# ---------------------------------------------------------------------------
# Detail-sheet reviewer-experience surfaces
# ---------------------------------------------------------------------------

def test_detail_sheet_has_frozen_header_and_join_key(tmp_path: Path):
    """``B2`` freezes both the header row and column A (Word ID).
    A reviewer paging through hundreds of rows still sees the column
    titles and the join key for the row they're on."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    assert ws.freeze_panes == "B2", (
        f"detail sheet freeze_panes={ws.freeze_panes!r}; expected 'B2'"
    )


def test_detail_sheet_has_autofilter_over_full_data_range(tmp_path: Path):
    """Autofilter is the cheap reviewer affordance: filter to LOW or
    UNRESOLVED rows without typing a formula. The range must cover the
    whole sheet so a sort-by-Confidence reorders every data row, not a
    truncated subset."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    from openpyxl.utils import get_column_letter
    expected = f"A1:{get_column_letter(len(REVIEW_HEADERS))}{ws.max_row}"
    assert ws.auto_filter.ref == expected, (
        f"detail sheet auto_filter.ref={ws.auto_filter.ref!r}; "
        f"expected {expected!r}"
    )


def test_detail_sheet_has_sensible_column_widths(tmp_path: Path):
    """Without explicit widths, openpyxl writes default-width columns
    and the reviewer has to manually resize every column. Pin a
    positive lower bound on the four trust-slice columns (Word ID,
    Location, Snippet, Raw Token) so a regression that drops the
    width settings is caught."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    for col_letter in ("A", "B", "C", "D"):
        width = ws.column_dimensions[col_letter].width
        assert width is not None and width >= 10, (
            f"column {col_letter} width={width!r}; reviewer-friendly "
            "widths must be set on the trust-slice columns"
        )


def test_detail_sheet_paints_confidence_column_per_status(tmp_path: Path):
    """Color fills on the Confidence column let a reviewer scan the
    sheet visually and immediately tell which rows still need
    attention. Every observed confidence label must paint with the
    matching ``CONFIDENCE_FILLS`` entry."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    conf_col = REVIEW_HEADERS.index("Confidence") + 1
    seen_confidences: set = set()
    for r in range(2, ws.max_row + 1):
        conf = ws.cell(row=r, column=conf_col).value
        seen_confidences.add(conf)
        cell_fill = ws.cell(row=r, column=conf_col).fill
        expected = CONFIDENCE_FILLS.get(conf)
        if expected is None:
            continue
        # PatternFill carries fgColor.rgb; compare on the rgb string
        # because openpyxl wraps Color objects on read and identity
        # comparison would false-fail.
        assert (cell_fill.fgColor.rgb or "").upper().endswith(
            expected.fgColor.rgb.upper()
        ), (
            f"row {r} confidence={conf} has fill rgb="
            f"{cell_fill.fgColor.rgb!r}; expected ending in "
            f"{expected.fgColor.rgb!r}"
        )
    # Synthetic corpus has at least HIGH + EXCLUDED rows; pin a non-empty
    # observed-confidences set so a regression that empties the column
    # fails here loudly.
    assert seen_confidences, "detail sheet had no confidence values to color"


def test_detail_sheet_paints_review_status_column_per_status(tmp_path: Path):
    """Same color story for the Review Status column — the trust-slice
    audit field that tells a reviewer where the row stands in the
    confirm-mapping pipeline."""
    xlsx = _run_learn(tmp_path)
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["mapping_review"]
    review_col = REVIEW_HEADERS.index("Review Status") + 1
    for r in range(2, ws.max_row + 1):
        review = ws.cell(row=r, column=review_col).value
        expected = REVIEW_STATUS_FILLS.get(review)
        if expected is None:
            continue
        cell_fill = ws.cell(row=r, column=review_col).fill
        assert (cell_fill.fgColor.rgb or "").upper().endswith(
            expected.fgColor.rgb.upper()
        ), (
            f"row {r} review_status={review} has fill rgb="
            f"{cell_fill.fgColor.rgb!r}; expected ending in "
            f"{expected.fgColor.rgb!r}"
        )
