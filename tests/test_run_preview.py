"""Tests for the run-mode preview gate (``run-preview`` CLI / ``src.run_preview``).

Contracts under test (mirrored from src.run_preview's docstring):

* A successful run resolves every confirmed entry against the new Excel
  workbook, applies the recorded transform deterministically, and writes
  ``run_validation.xlsx`` with one row per confirmed ``word_id``.
* Missing source sheet, missing/empty source cell, non-numeric source
  cell, and unknown transform interpretations all fail loudly — the row
  is surfaced in the artifact with a per-row status and the CLI exits 7.
* An incomplete ``confirmed_mapping.yml`` (any ``review_required`` row,
  ``summary.complete: false``, or ``summary.allow_incomplete: true``) is
  refused before any extraction work — the CLI exits 6 and no artifact
  is written.
* No silent omission: every confirmed entry shows up in the artifact and
  in the in-memory report, even when its row failed resolution.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pytest
import yaml

from src.main import main as cli_main
from src.run_preview import (
    STATUS_MISSING_CELL,
    STATUS_MISSING_SHEET,
    STATUS_MISSING_TRANSFORM,
    STATUS_NON_NUMERIC,
    STATUS_OK,
    STATUS_TRANSFORM_UNKNOWN,
    run_preview,
    write_run_validation,
)


# ---------------------------------------------------------------------------
# Tiny fixture builders — a fake new-period Excel + a fake confirmed YAML
# ---------------------------------------------------------------------------

def _write_excel(path: Path, sheets: Dict[str, List[List]]) -> Path:
    """Write a workbook with the given ``{sheet_name: [[row]...]}`` shape."""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for r in rows:
            ws.append(r)
    wb.save(str(path))
    return path


def _confirmed_entry(
    word_id: str,
    *,
    sheet: str,
    cell: str,
    value: float,
    interpretation: str = "as_written",
    raw: str = "100",
    word_value: float = 100.0,
    unit: str = "",
    confidence: str = "HIGH",
    snippet: str = "100 收入",
    location: str = "paragraph:0",
) -> Dict:
    """Build one ``confirmed_mappings`` entry that resembles real output."""
    return {
        "word_id": word_id,
        "location": location,
        "raw": raw,
        "value": word_value,
        "unit": unit,
        "sign": 1,
        "status": confidence,
        "confidence": confidence,
        "review_status": "confirmed",
        "reviewer_decision": "confirm",
        "reviewer_notes": "",
        "context": {"snippet": snippet, "label_context": []},
        "recommended_source": {
            "sheet": sheet, "address": cell, "value": value,
            "row_context": ["收入"], "column_context": ["2026年6月"],
        },
        "confirmed_source": {
            "sheet": sheet, "address": cell, "value": value,
        },
        "source_origin": "recommended",
        "transform": {
            "interpretation": interpretation,
            "value_score": 1.0,
            "context_score": 1.0,
            "overlap_tokens": ["收入"],
        },
        "placeholder": "{{ " + word_id + " }}",
        "placeholder_status": "applied",
    }


def _write_confirmed_yaml(
    path: Path,
    confirmed: List[Dict],
    *,
    review_required: Optional[List[Dict]] = None,
    audit_only: Optional[List[Dict]] = None,
    complete: bool = True,
    allow_incomplete: bool = False,
) -> Path:
    review_required = review_required or []
    audit_only = audit_only or []
    doc = {
        "schema_version": 1,
        "source_artifacts": {
            "auto_mapping": "fixture/auto_mapping.yml",
            "mapping_review": "fixture/mapping_review.xlsx",
        },
        "summary": {
            "total_word_numbers": (
                len(confirmed) + len(review_required) + len(audit_only)
            ),
            "confirmed": len(confirmed),
            "review_required": len(review_required),
            "audit_only_excluded": len(audit_only),
            "allow_incomplete": allow_incomplete,
            "complete": complete,
        },
        "confirmed_mappings": confirmed,
        "review_required": review_required,
        "audit_only_excluded": audit_only,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump(doc, allow_unicode=True, sort_keys=False, width=1000),
        encoding="utf-8",
    )
    return path


def _read_artifact(path: Path) -> List[Dict]:
    """Read ``run_validation.xlsx`` back as a list of header-keyed dicts."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out: List[Dict] = []
    for r in rows[1:]:
        out.append({h: v for h, v in zip(header, r)})
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Successful extraction
# ---------------------------------------------------------------------------

def test_run_preview_successful_extraction(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度核心指标": [
            ["指标", "2026年6月"],
            ["营业收入(元)", 234_567_890.0],
            ["毛利率(%)", 37.5],
        ],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry(
            "word_0001", sheet="月度核心指标", cell="B2",
            value=234_567_890.0,
            interpretation="万元→base_unit",
            raw="23,456.79万元", word_value=23456.79, unit="万元",
            snippet="6月营业收入达23,456.79万元",
        ),
        _confirmed_entry(
            "word_0002", sheet="月度核心指标", cell="B3", value=37.5,
            interpretation="as_written",
            raw="37.50%", word_value=37.5, unit="%",
            snippet="毛利率达37.50%",
        ),
    ])

    report = run_preview(excel, confirmed)

    assert report.fatal_errors == []
    assert len(report.rows) == 2
    assert all(r.status == STATUS_OK for r in report.rows), [
        (r.word_id, r.status, r.detail) for r in report.rows
    ]
    by_id = {r.word_id: r for r in report.rows}
    # 万元→base_unit row: 234,567,890 / 10,000 = 23,456.789 generated.
    assert by_id["word_0001"].raw_excel_value == 234_567_890.0
    assert by_id["word_0001"].generated_value == pytest.approx(23456.789)
    assert by_id["word_0001"].transform_interpretation == "万元→base_unit"
    # as_written row: pass-through.
    assert by_id["word_0002"].raw_excel_value == pytest.approx(37.5)
    assert by_id["word_0002"].generated_value == pytest.approx(37.5)


def test_run_preview_applies_percent_transform(tmp_path: Path):
    # %→decimal: Word writes "12.34%" with raw value 12.34; Excel stores
    # 0.1234. generated = excel / (1/100) = excel * 100 = 12.34.
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["转化率", 0.0567]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry(
            "word_0001", sheet="月度", cell="B2", value=0.0567,
            interpretation="%→decimal",
            raw="5.67%", word_value=5.67, unit="%",
        ),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_OK
    assert row.generated_value == pytest.approx(5.67)


def test_run_preview_applies_yi_yuan_transform(tmp_path: Path):
    # 亿元→base_unit: Excel holds 1,234,567,890; Word writes "12.35亿元"
    # with raw value 12.35. generated = excel / 1e8 ≈ 12.3456789.
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 1_234_567_890.0]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry(
            "word_0001", sheet="月度", cell="B2", value=1_234_567_890.0,
            interpretation="亿元→base_unit",
            raw="12.35亿元", word_value=12.35, unit="亿元",
        ),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_OK
    assert row.generated_value == pytest.approx(12.3456789)


# ---------------------------------------------------------------------------
# Fail-loud per-row failures
# ---------------------------------------------------------------------------

def test_run_preview_fails_when_source_sheet_missing(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "Other Sheet": [["x", 1], ["y", 2]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0),
    ])
    report = run_preview(excel, confirmed)
    assert report.fatal_errors == []
    [row] = report.rows
    assert row.status == STATUS_MISSING_SHEET
    assert "月度" in row.detail


def test_run_preview_fails_when_source_cell_empty(tmp_path: Path):
    # Sheet exists, but the targeted cell is empty in the new workbook.
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["指标"]]})
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_MISSING_CELL


def test_run_preview_fails_when_cell_non_numeric(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "value"], ["营收", "N/A"]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=0.0),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_NON_NUMERIC
    assert "N/A" in row.detail


def test_run_preview_fails_when_cell_holds_bool(tmp_path: Path):
    # ``True`` is a subclass of int in Python; it must NOT be silently
    # treated as 1.0 for run-preview rendering.
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "value"], ["营收", True]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=1.0),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_NON_NUMERIC


def test_run_preview_fails_on_unknown_transform(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "value"], ["营收", 100.0]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry(
            "word_0001", sheet="月度", cell="B2", value=100.0,
            interpretation="bogus_v999",
        ),
    ])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_TRANSFORM_UNKNOWN
    assert "bogus_v999" in row.detail


def test_run_preview_fails_when_transform_missing(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "value"], ["营收", 100.0]],
    })
    entry = _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0)
    entry["transform"] = None
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [entry])
    [row] = run_preview(excel, confirmed).rows
    assert row.status == STATUS_MISSING_TRANSFORM


# ---------------------------------------------------------------------------
# Incomplete-mapping refusal
# ---------------------------------------------------------------------------

def test_run_preview_refuses_review_required_present(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    confirmed = _write_confirmed_yaml(
        tmp_path / "confirmed.yml",
        [_confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)],
        review_required=[{"word_id": "word_0002", "reason": "blank_decision"}],
        complete=False,
    )
    report = run_preview(excel, confirmed)
    assert report.fatal_errors, "incomplete mapping must produce a fatal error"
    assert any("review_required" in e for e in report.fatal_errors)
    # No extraction work attempted when the gate refuses.
    assert report.rows == []


def test_run_preview_refuses_summary_complete_false(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    confirmed = _write_confirmed_yaml(
        tmp_path / "confirmed.yml",
        [_confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)],
        complete=False,
    )
    report = run_preview(excel, confirmed)
    assert report.fatal_errors


def test_run_preview_refuses_allow_incomplete(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    confirmed = _write_confirmed_yaml(
        tmp_path / "confirmed.yml",
        [_confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)],
        complete=False, allow_incomplete=True,
    )
    report = run_preview(excel, confirmed)
    assert report.fatal_errors
    assert any("allow_incomplete" in e for e in report.fatal_errors)


def test_run_preview_refuses_empty_confirmed_mapping(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    confirmed = _write_confirmed_yaml(
        tmp_path / "confirmed.yml", [], complete=True,
    )
    report = run_preview(excel, confirmed)
    assert report.fatal_errors


# ---------------------------------------------------------------------------
# Fail-CLOSED completeness gate: missing/wrong-shape metadata must refuse,
# not silently fall through the `is False` / `is True` checks.
# ---------------------------------------------------------------------------

def _write_raw_yaml(path: Path, doc: Dict) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.safe_dump(doc, allow_unicode=True, sort_keys=False, width=1000),
        encoding="utf-8",
    )
    return path


def test_run_preview_refuses_missing_summary_block(tmp_path: Path):
    """A confirmed_mapping.yml with no summary block can't prove
    completeness — the gate must refuse rather than silently use the
    file just because ``confirmed_mappings`` happens to look populated.
    """
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    entry = _confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)
    confirmed = _write_raw_yaml(tmp_path / "confirmed.yml", {
        "schema_version": 1,
        "confirmed_mappings": [entry],
    })
    report = run_preview(excel, confirmed)
    assert report.fatal_errors
    assert any("summary" in e for e in report.fatal_errors)
    assert report.rows == []


def test_run_preview_refuses_summary_missing_complete_field(tmp_path: Path):
    """If the summary block exists but ``complete`` is absent, the gate
    must NOT silently pass — the previous ``is False`` check would have
    let ``None`` through. Fail-closed: only the literal ``True`` proves
    completeness."""
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    entry = _confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)
    confirmed = _write_raw_yaml(tmp_path / "confirmed.yml", {
        "schema_version": 1,
        "summary": {
            "total_word_numbers": 1,
            "confirmed": 1,
            "review_required": 0,
            "audit_only_excluded": 0,
            "allow_incomplete": False,
            # complete: missing on purpose.
        },
        "confirmed_mappings": [entry],
        "review_required": [],
        "audit_only_excluded": [],
    })
    report = run_preview(excel, confirmed)
    assert report.fatal_errors
    assert any("complete" in e for e in report.fatal_errors)
    assert report.rows == []


def test_run_preview_refuses_complete_none(tmp_path: Path):
    """``complete: null`` (YAML None) must NOT pass the gate."""
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    entry = _confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)
    confirmed = _write_raw_yaml(tmp_path / "confirmed.yml", {
        "schema_version": 1,
        "summary": {
            "total_word_numbers": 1,
            "confirmed": 1,
            "review_required": 0,
            "audit_only_excluded": 0,
            "allow_incomplete": False,
            "complete": None,
        },
        "confirmed_mappings": [entry],
        "review_required": [],
        "audit_only_excluded": [],
    })
    report = run_preview(excel, confirmed)
    assert report.fatal_errors
    assert any("complete" in e for e in report.fatal_errors)


def test_run_preview_refuses_complete_truthy_but_not_true(tmp_path: Path):
    """A near-miss like ``complete: "true"`` (a string) or
    ``complete: 1`` is *not* the boolean ``True`` — fail-closed."""
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    for bad in ("true", "yes", 1, "True"):
        entry = _confirmed_entry("word_0001", sheet="月度", cell="B1", value=1.0)
        confirmed = _write_raw_yaml(tmp_path / f"confirmed_{bad}.yml", {
            "schema_version": 1,
            "summary": {
                "total_word_numbers": 1,
                "confirmed": 1,
                "review_required": 0,
                "audit_only_excluded": 0,
                "allow_incomplete": False,
                "complete": bad,
            },
            "confirmed_mappings": [entry],
            "review_required": [],
            "audit_only_excluded": [],
        })
        report = run_preview(excel, confirmed)
        assert report.fatal_errors, f"value {bad!r} must not pass the gate"
        assert any("complete" in e for e in report.fatal_errors)


def test_run_preview_accepts_only_explicit_true_complete(tmp_path: Path):
    """Positive control: the same fixture with ``complete: true`` proceeds
    past the gate. Pins the boundary between accept and refuse."""
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 100.0]],
    })
    entry = _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0)
    confirmed = _write_raw_yaml(tmp_path / "confirmed.yml", {
        "schema_version": 1,
        "summary": {
            "total_word_numbers": 1,
            "confirmed": 1,
            "review_required": 0,
            "audit_only_excluded": 0,
            "allow_incomplete": False,
            "complete": True,
        },
        "confirmed_mappings": [entry],
        "review_required": [],
        "audit_only_excluded": [],
    })
    report = run_preview(excel, confirmed)
    assert report.fatal_errors == []
    [row] = report.rows
    assert row.status == STATUS_OK


# ---------------------------------------------------------------------------
# No-silent-omission invariant
# ---------------------------------------------------------------------------

def test_run_preview_includes_every_confirmed_word_id(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 100.0], ["毛利", 200.0]],
        "周度": [["指标", "v"], ["GMV", 300_000.0]],
    })
    entries = [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0),
        _confirmed_entry("word_0002", sheet="月度", cell="B3", value=200.0),
        _confirmed_entry(
            "word_0003", sheet="周度", cell="B2", value=300_000.0,
            interpretation="万元→base_unit",
        ),
    ]
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", entries)
    report = run_preview(excel, confirmed)
    seen = {r.word_id for r in report.rows}
    assert seen == {e["word_id"] for e in entries}


def test_run_preview_artifact_lists_every_confirmed_word_id_even_on_failure(
    tmp_path: Path,
):
    """Even when one row fails, the artifact still has a row for every
    confirmed ``word_id`` — that's the audit guarantee CLAUDE.md mandates.
    """
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 100.0]],
    })
    entries = [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0),
        # word_0002 points at a cell that doesn't exist in the new excel.
        _confirmed_entry("word_0002", sheet="月度", cell="Z99", value=999.0),
        # word_0003 points at a sheet that doesn't exist.
        _confirmed_entry("word_0003", sheet="周度", cell="B2", value=1.0),
    ]
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", entries)
    report = run_preview(excel, confirmed)
    out = write_run_validation(report, tmp_path / "out")
    rows = _read_artifact(out)
    assert [r["Word ID"] for r in rows] == ["word_0001", "word_0002", "word_0003"]
    by_id = {r["Word ID"]: r for r in rows}
    assert by_id["word_0001"]["Status"] == STATUS_OK
    assert by_id["word_0002"]["Status"] == STATUS_MISSING_CELL
    assert by_id["word_0003"]["Status"] == STATUS_MISSING_SHEET


# ---------------------------------------------------------------------------
# Artifact schema (column coverage)
# ---------------------------------------------------------------------------

def test_run_validation_artifact_carries_required_columns(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 100.0]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry(
            "word_0001", sheet="月度", cell="B2", value=100.0,
            raw="100元", word_value=100.0, unit="元",
            snippet="本月营收达100元",
        ),
    ])
    report = run_preview(excel, confirmed)
    out = write_run_validation(report, tmp_path / "out")
    [row] = _read_artifact(out)
    # The goal pinned these as the row-level contract for run-preview.
    for required in (
        "Word ID", "Word Location", "Word Context", "Word Raw Token",
        "Source Sheet", "Source Cell", "Raw Excel Value",
        "Generated Value", "Transform Interpretation",
        "Confidence", "Status",
    ):
        assert required in row, f"artifact missing required column {required!r}"
    assert row["Source Sheet"] == "月度"
    assert row["Source Cell"] == "B2"
    assert row["Confidence"] == "HIGH"
    assert row["Word Context"] == "本月营收达100元"


# ---------------------------------------------------------------------------
# CLI exit codes
# ---------------------------------------------------------------------------

def test_cli_run_preview_success_writes_artifact_and_returns_0(
    tmp_path: Path, capsys,
):
    excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["指标", "v"], ["营收", 100.0]],
    })
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B2", value=100.0),
    ])
    out = tmp_path / "preview_out"
    rc = cli_main([
        "run-preview",
        "--excel", str(excel),
        "--confirmed", str(confirmed),
        "--out", str(out),
    ])
    assert rc == 0
    artifact = out / "run_validation.xlsx"
    assert artifact.exists()
    captured = capsys.readouterr()
    assert "run-preview summary" in captured.out
    assert "ok" in captured.out


def test_cli_run_preview_missing_cell_returns_7_with_artifact(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x"]]})
    confirmed = _write_confirmed_yaml(tmp_path / "confirmed.yml", [
        _confirmed_entry("word_0001", sheet="月度", cell="B99", value=1.0),
    ])
    out = tmp_path / "preview_out"
    rc = cli_main([
        "run-preview",
        "--excel", str(excel),
        "--confirmed", str(confirmed),
        "--out", str(out),
    ])
    assert rc == 7, "any per-row failure must produce a non-zero exit"
    # Artifact still written so the reviewer can see exactly which row broke.
    assert (out / "run_validation.xlsx").exists()


def test_cli_run_preview_incomplete_returns_6_no_artifact(tmp_path: Path):
    excel = _write_excel(tmp_path / "new.xlsx", {"月度": [["x", 1]]})
    confirmed = _write_confirmed_yaml(
        tmp_path / "confirmed.yml", [],
        review_required=[{"word_id": "word_0001", "reason": "blank_decision"}],
        complete=False,
    )
    out = tmp_path / "preview_out"
    rc = cli_main([
        "run-preview",
        "--excel", str(excel),
        "--confirmed", str(confirmed),
        "--out", str(out),
    ])
    assert rc == 6, "incomplete confirmed_mapping.yml must refuse to run"
    # An incomplete confirmed mapping is unusable — nothing to write.
    assert not (out / "run_validation.xlsx").exists()


def test_cli_run_preview_missing_inputs_returns_2(tmp_path: Path, capsys):
    rc = cli_main([
        "run-preview",
        "--excel", str(tmp_path / "nope.xlsx"),
        "--confirmed", str(tmp_path / "nope.yml"),
        "--out", str(tmp_path / "out"),
    ])
    assert rc == 2
    err = capsys.readouterr().err
    assert "not found" in err


# ---------------------------------------------------------------------------
# End-to-end alternative-override path: confirm-mapping → run-preview
# ---------------------------------------------------------------------------

def test_alternative_override_applies_correct_transform_end_to_end(
    tmp_path: Path,
):
    """confirm-mapping must hand run-preview the *alternative*'s
    transform, not the recommended pick's, when the reviewer overrode
    to an alternative. Otherwise run-preview applies the wrong factor
    silently and prints a confidently-wrong ``Generated Value``.

    Setup: recommended pick is at ``月度!B2`` with interpretation
    ``万元→base_unit`` (Excel ÷ 10,000). The reviewer overrides to an
    alternative at ``周度!B6`` whose own interpretation is ``as_written``.
    If confirm-mapping leaks the recommended transform onto the
    alternative, run-preview will divide 100 by 10,000 and emit 0.01 —
    instead of the correct 100. The test pins the correct value.
    """
    import openpyxl as _openpyxl

    from src.mapping_confirmer import confirm_mappings, write_confirmed_yaml
    from src.mapping_reviewer import REVIEW_HEADERS

    # --- Build a fake auto_mapping.yml with one HIGH entry whose
    # recommended pick uses a different interpretation than the alt. ---
    rec = {
        "sheet": "月度", "address": "B2", "value": 1_000_000.0,
        "row_context": ["营收(元)"], "column_context": ["2026年6月"],
    }
    alt = {
        "sheet": "周度", "address": "B6", "value": 100.0,
        "row_context": ["GMV(万元)"], "column_context": ["第21周"],
        "interpretation": "as_written",
        "value_score": 1.0, "context_score": 0.85,
        "overlap_tokens": ["GMV"],
    }
    auto_entry = {
        "word_id": "word_0001",
        "location": "paragraph:0",
        "raw": "100",
        "value": 100.0,
        "unit": "",
        "sign": 1,
        "status": "HIGH",
        "confidence": "HIGH",
        "review_status": "pending_review",
        "note": "",
        "placeholder": "{{ word_0001 }}",
        "placeholder_status": "applied",
        "context": {"snippet": "本周GMV为100", "label_context": []},
        "recommended_source": rec,
        "transform": {
            # Deliberately the WRONG factor for the alt cell.
            "interpretation": "万元→base_unit",
            "value_score": 0.85, "context_score": 1.0,
            "overlap_tokens": ["营收"],
        },
        "alternatives": [alt],
    }
    auto_path = tmp_path / "auto_mapping.yml"
    auto_path.write_text(yaml.safe_dump({
        "schema_version": 1,
        "summary": {
            "total": 1,
            "by_confidence": {
                "HIGH": 1, "MEDIUM": 0, "LOW": 0,
                "UNRESOLVED": 0, "EXCLUDED": 0,
            },
            "placeholders_applied": 1,
        },
        "mappings": [auto_entry],
    }, allow_unicode=True, sort_keys=False, width=1000), encoding="utf-8")

    # --- Build the mapping_review.xlsx with a reviewer override to the alt. ---
    review_path = tmp_path / "mapping_review.xlsx"
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "mapping_review"
    ws.append(REVIEW_HEADERS)
    col_idx = {h: REVIEW_HEADERS.index(h) for h in REVIEW_HEADERS}
    row = [""] * len(REVIEW_HEADERS)
    row[col_idx["Word ID"]] = "word_0001"
    row[col_idx["Word Location"]] = "paragraph:0"
    row[col_idx["Word Raw Token"]] = "100"
    row[col_idx["Confidence"]] = "HIGH"
    row[col_idx["Review Status"]] = "pending_review"
    row[col_idx["Placeholder Status"]] = "applied"
    row[col_idx["Placeholder"]] = "{{ word_0001 }}"
    row[col_idx["Top Excel Sheet"]] = "月度"
    row[col_idx["Top Excel Cell"]] = "B2"
    row[col_idx["Reviewer Decision"]] = "confirm"
    row[col_idx["Confirmed Sheet"]] = "周度"
    row[col_idx["Confirmed Cell"]] = "B6"
    ws.append(row)
    wb.save(str(review_path))

    # --- Run confirm-mapping → confirmed_mapping.yml. ---
    confirmed_path = tmp_path / "confirmed_mapping.yml"
    report = confirm_mappings(auto_path, review_path)
    write_confirmed_yaml(
        report, auto_path, review_path, confirmed_path,
        allow_incomplete=False, total_word_numbers=1,
    )

    # Sanity: the confirmed entry carries the alternative's transform.
    confirmed_doc = yaml.safe_load(confirmed_path.read_text(encoding="utf-8"))
    [conf] = confirmed_doc["confirmed_mappings"]
    assert conf["source_origin"] == "reviewer_override:alternative"
    assert conf["transform"]["interpretation"] == "as_written", (
        "alternative override must carry the alternative's own "
        "interpretation, not the recommended pick's"
    )

    # --- Build a NEW excel where 周度!B6 = 100, and run preview. ---
    new_excel = _write_excel(tmp_path / "new.xlsx", {
        "月度": [["x", "y"], ["营收", 999_999_999.0]],
        "周度": [["a", "b"], ["c", "d"], ["e", "f"], ["g", "h"],
                 ["i", "j"], ["k", 100.0]],  # B6 = 100
    })
    preview_report = run_preview(new_excel, confirmed_path)
    [preview_row] = preview_report.rows
    assert preview_row.status == STATUS_OK
    assert preview_row.source_sheet == "周度"
    assert preview_row.source_cell == "B6"
    assert preview_row.raw_excel_value == 100.0
    # The correctness pin: with the right (as_written) transform,
    # generated == 100. With the bug (万元→base_unit), it would be 0.01.
    assert preview_row.generated_value == pytest.approx(100.0), (
        f"alternative override applied wrong factor: got "
        f"{preview_row.generated_value!r}; expected 100.0. "
        "confirm-mapping is leaking the recommended pick's transform "
        "onto an alternative override."
    )
